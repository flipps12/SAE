from django.views.generic import ListView, FormView, UpdateView, TemplateView
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Prefetch, Count
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin

from gestion_academica.models import Persona, Alumno, Curso, CicloLectivo, HistorialAcademico, Especialidad, InscripcionDictado, Dictado
from gestion_academica.forms import AlumnoForm, PersonaForm, ImportarAlumnosForm, AlumnoFormSet


import random

User = get_user_model()


class ListadoAlumnosView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Alumno
    template_name = 'alumnos/listado_alumnos.html'
    context_object_name = 'alumnos'
    paginate_by = 25  

    # Configuración de restricciones:
    # 1. Obliga a que el usuario pertenezca al Staff (is_staff=True)
    staff_member_required = True 
    
    # 2. Si un usuario logueado NO es staff, Django le tira un error 403 Forbidden automático.
    # Si preferís que los redirija al Home en vez de tirar el error, descomentá estas dos líneas:
    # raise_exception = False
    # login_url = 'home'

    def has_permission(self):
        """
        Este método define la regla de entrada. 
        Solo permite el acceso si el usuario está activo y es staff.
        """
        return self.request.user.is_active and self.request.user.is_staff

    def get_queryset(self):
        # Traemos la persona de antemano y los historiales con sus respectivos cursos
        queryset = Alumno.objects.select_related('persona').prefetch_related(
            'historiales__curso__especialidad', 
            'historiales__ciclo_lectivo'
        )

        # Captura de parámetros del buscador y los selectores
        search_query = self.request.GET.get('q', '').strip()
        curso_filter = self.request.GET.get('curso', '')
        especialidad_filter = self.request.GET.get('especialidad', '')
        estado_filter = self.request.GET.get('estado', '')

        # Filtro 1: Buscador General (Nombre, Apellido, DNI, Número de Legajo)
        if search_query:
            queryset = queryset.filter(
                Q(persona__apellido__icontains=search_query) |
                Q(persona__nombre__icontains=search_query) |
                Q(persona__dni__icontains=search_query) |
                Q(persona__numero_legajo__icontains=search_query)
            )

        # Filtro 2: Por Curso (mediante Historial Académico)
        if curso_filter:
            queryset = queryset.filter(historiales__curso_id=curso_filter)

        # Filtro 3: Por Especialidad (mediante Historial Académico)
        if cuestion_filter := especialidad_filter:
            queryset = queryset.filter(historiales__curso__especialidad_id=cuestion_filter)

        # Filtro 4: Por Estado de Actividad de la Institución
        if estado_filter == 'activos':
            queryset = queryset.filter(activo=True)
        elif estado_filter == 'inactivos':
            queryset = queryset.filter(activo=False)

        # Retornamos el queryset limpio quitando duplicados si un alumno tiene múltiples historiales
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Contexto necesario para renderizar los elementos select de los filtros
        context['cursos'] = Curso.objects.all()
        context['especialidades'] = Especialidad.objects.all()
        
        # Mantener los valores actuales en los campos del formulario tras recargar
        context['search_query'] = self.request.GET.get('q', '')
        context['curso_seleccionado'] = self.request.GET.get('curso', '')
        context['especialidad_seleccionada'] = self.request.GET.get('especialidad', '')
        context['estado_seleccionado'] = self.request.GET.get('estado', '')
        
        return context


class ImportarAlumnosView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'alumnos/importar_alumnos.html'
    form_class = ImportarAlumnosForm

    def test_func(self):
        """
        Esta función define el filtro de seguridad de Django:
        Retorna True solo si el usuario está logueado y es staff (is_staff=True).
        """
        return self.request.user.is_staff

    def handle_no_permission(self):
        """
        Si el usuario no es staff, le muestra un mensaje de error y
        lo redirige, evitando que vea la pantalla o pueda subir archivos.
        """
        messages.error(self.request, "Acceso denegado. Solo el personal jerárquico o administrativo (Staff) puede realizar importaciones masivas.")
        return super().handle_no_permission()

    def form_valid(self, form):
        # Capturamos el archivo subido
        archivo = self.request.FILES['archivo_excel']
        resultados = None

        try:
            # Leer el archivo directo de la memoria con openpyxl
            wb = load_workbook(filename=archivo, data_only=True)
            sheet = wb.active
            
            # Mapeamos qué columna tiene cada dato según la fila 1 (Cabecera)
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1) if cell.value is not None}
            
            # Columnas obligatorias que vamos a exigir
            required_headers = [
                'username', 'password', 'nombre', 'apellido', 'dni', 'cuil', 
                'numero_legajo', 'fecha_nacimiento', 'libro', 'folio', 'activo', 
                'ciclo_lectivo', 'curso_nivel', 'curso_division', 'especialidad'
            ]
            
            # Si falta alguna columna, frenamos el proceso de inmediato
            faltantes = [req for req in required_headers if req not in headers]
            if faltantes:
                messages.error(self.request, f"Al Excel le faltan columnas críticas: {', '.join(faltantes)}")
                return self.form_invalid(form)

            exitosos = []
            errores = []

            # Recorremos desde la fila 2 en adelante
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                # Si la celda de usuario está vacía, ignoramos la fila de manera segura
                if row[headers['username'] - 1].value is None:
                    continue
                
                # Armamos un diccionario con los datos limpios de la fila actual
                data = {key: row[idx - 1].value for key, idx in headers.items()}
                alumno_str = f"{data.get('apellido', '')}, {data.get('nombre', '')} (DNI: {data.get('dni', '')})"

                try:
                    # Bloque aislado por fila: si esta falla, no afecta a los demás alumnos
                    with transaction.atomic():
                        
                        # 1. Validar infraestructura
                        ciclo = CicloLectivo.objects.get(anio=int(data['ciclo_lectivo']))
                        curso = Curso.objects.get(
                            nivel=int(data['curso_nivel']),
                            division=int(data['curso_division']),
                            especialidad__nombre__iexact=str(data['especialidad']).strip()
                        )

                        # 2. Crear credenciales de acceso
                        user, user_created = User.objects.get_or_create(
                            username=str(data['username']).strip()
                        )
                        if user_created:
                            user.set_password(str(data['password']))
                            user.save()

                        # 3. Crear datos personales base
                        persona, _ = Persona.objects.get_or_create(
                            dni=str(data['dni']).strip(),
                            defaults={
                                'user': user,
                                'nombre': str(data['nombre']).strip(),
                                'apellido': str(data['apellido']).strip(),
                                'cuil': str(data['cuil']).strip(),
                                'numero_legajo': str(data['numero_legajo']).strip(),
                                'fecha_nacimiento': data['fecha_nacimiento']
                            }
                        )

                        # 4. Crear el perfil de Alumno
                        activo_bool = True if str(data['activo']).upper() in ['TRUE', '1', 'SÍ', 'SI'] else False
                        alumno, _ = Alumno.objects.get_or_create(
                            persona=persona,
                            defaults={
                                'libro': str(data['libro'] or '').strip(),
                                'folio': str(data['folio'] or '').strip(),
                                'activo': activo_bool
                            }
                        )

                        # 5. Generar la inscripción (Historial Académico)
                        _, hist_created = HistorialAcademico.objects.get_or_create(
                            alumno=alumno,
                            ciclo_lectivo=ciclo,
                            defaults={'curso': curso, 'estado_final': 'CURSANDO'}
                        )

                        msg = f"Inscripto con éxito en {curso}" if hist_created else "El alumno ya pertenecía a este ciclo."
                        exitosos.append({'fila': row_idx, 'alumno': alumno_str, 'detalle': msg})

                except CicloLectivo.DoesNotExist:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': f"El Ciclo Lectivo {data['ciclo_lectivo']} no existe."})
                except Curso.DoesNotExist:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': f"El Curso {data['curso_nivel']}° {data['curso_division']}ª ({data['especialidad']}) no existe."})
                except Exception as e:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': str(e)})

            resultados = {'exitosos': exitosos, 'errores': errores}
            messages.success(self.request, f"Procesamiento terminado. Éxitos: {len(exitosos)} | Errores: {len(errores)}")

        except Exception as e:
            messages.error(self.request, f"Error al procesar la estructura del Excel: {e}")

        # Retornamos la misma pantalla inyectando el diccionario de resultados para renderizar las alertas
        return render(self.request, self.template_name, {'form': form, 'resultados': resultados})


class EditarAlumnoView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'alumnos/editar_alumno.html'
    success_url = reverse_lazy('listado_alumnos')

    def has_permission(self):
        return self.request.user.is_active and self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()
        context['ciclo_activo'] = ciclo_activo
        
        if ciclo_activo:
            # Traemos las materias actuales ordenadas por nombre
            context['inscripciones'] = InscripcionDictado.objects.filter(
                alumno=self.object, 
                ciclo_lectivo=ciclo_activo
            ).select_related('dictado__materia', 'dictado__curso')
            
            # Traemos TODOS los dictados del año para el selector de "Agregar Materia"
            context['todos_los_dictados'] = Dictado.objects.filter(
                ciclo_lectivo=ciclo_activo
            ).select_related('materia', 'curso')

        if self.request.POST:
            context['persona_form'] = PersonaForm(self.request.POST, instance=self.object.persona)
        else:
            context['persona_form'] = PersonaForm(instance=self.object.persona)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object_with_id() # Obtenemos el alumno actual
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()

        # ACCIÓN A: Eliminar una materia suelta
        if 'action_eliminar_materia' in request.POST:
            inscripcion_id = request.POST.get('inscripcion_id')
            inscripcion = get_object_or_404(InscripcionDictado, id=inscripcion_id, alumno=self.object)
            materia_nombre = inscripcion.dictado.materia.nombre
            inscripcion.delete()
            messages.warning(request, f"Se eliminó la inscripción a la materia {materia_nombre}.")
            return render(request, self.template_name, self.get_context_data())

        # ACCIÓN B: Agregar una materia suelta (Recursante / Previa)
        if 'action_agregar_materia' in request.POST:
            dictado_id = request.POST.get('dictado_id')
            condicion = request.POST.get('condicion_materia', 'RECURSANTE')
            
            if dictado_id:
                dictado = get_object_or_404(Dictado, id=dictado_id, ciclo_lectivo=ciclo_activo)
                nueva_inscripcion, created = InscripcionDictado.objects.get_or_create(
                    alumno=self.object,
                    dictado=dictado,
                    ciclo_lectivo=ciclo_activo,
                    defaults={'condicion': condicion}
                )
                if created:
                    messages.success(request, f"Se agregó la materia {dictado.materia.nombre} como {nueva_inscripcion.get_condicion_display()}.")
                else:
                    messages.info(request, f"El alumno ya estaba inscripto en {dictado.materia.nombre}.")
            else:
                messages.error(request, "Debe seleccionar una materia válida de la lista.")
            return render(request, self.template_name, self.get_context_data())

        # ACCIÓN C: Guardado clásico de todo el formulario (Persona + Alumno)
        return super().post(request, *args, **kwargs)

    def get_object_with_id(self):
        # Auxiliar seguro para recuperar el alumno en el método post
        return get_object_or_404(Alumno, pk=self.kwargs.get(self.pk_url_kwarg or 'pk'))

    def form_valid(self, form):
        context = self.get_context_data()
        persona_form = context['persona_form']
        ciclo_activo = context['ciclo_activo']
        
        if form.is_valid() and persona_form.is_valid():
            if not ciclo_activo:
                messages.error(self.request, "No se puede guardar la inscripción porque no hay un Ciclo Lectivo activo.")
                return self.form_invalid(form)
                
            try:
                with transaction.atomic():
                    persona_form.save()
                    self.object = form.save()
                    
                    curso_seleccionado = form.cleaned_data['curso']
                    burbuja_seleccionada = form.cleaned_data['burbuja']
                    taller_seleccionado = form.cleaned_data['grupo_taller']

                    historial, _ = HistorialAcademico.objects.get_or_create(
                        alumno=self.object,
                        ciclo_lectivo=ciclo_activo,
                        defaults={'curso': curso_seleccionado}
                    )
                    historial.curso = curso_seleccionado
                    historial.burbuja = burbuja_seleccionada
                    historial.grupo_taller = taller_seleccionado
                    historial.full_clean()
                    historial.save()
                    
                    # Inscripción masiva automática (solo si no existen previamente)
                    dictados_del_curso = Dictado.objects.filter(curso=curso_seleccionado, ciclo_lectivo=ciclo_activo)
                    for dictado in dictados_del_curso:
                        InscripcionDictado.objects.get_or_create(
                            alumno=self.object,
                            dictado=dictado,
                            ciclo_lectivo=ciclo_activo,
                            defaults={'condicion': 'REGULAR'}
                        )
                    
                messages.success(self.request, "El alumno y su grilla de materias se actualizaron correctamente.")
                return super().form_valid(form)
                
            except Exception as e:
                messages.error(self.request, f"Error al guardar los datos: {str(e)}")
                return self.form_invalid(form)
        else:
            return self.form_invalid(form)


class CargaFormsetAlumnosView(TemplateView):
    template_name = 'alumnos/carga_dinamica_alumnos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ciclos'] = CicloLectivo.objects.filter(activo=True)
        context['cursos'] = Curso.objects.all()
        context['formset'] = AlumnoFormSet(prefix='alumnos')
        return context

    def post(self, request, *args, **kwargs):
        ciclo_id = request.POST.get('ciclo_lectivo')
        curso_id = request.POST.get('curso')
        
        if not ciclo_id or not curso_id:
            messages.error(request, "Falta seleccionar el Ciclo Lectivo o el Curso de destino.")
            return self.render_to_response(self.get_context_data())

        ciclo = CicloLectivo.objects.get(id=ciclo_id)
        curso_destino = Curso.objects.get(id=curso_id)
        formset = AlumnoFormSet(request.POST, prefix='alumnos')

        if formset.is_valid():
            # Traer los dictados asignados a este curso y ciclo
            dictados_curso = Dictado.objects.filter(curso=curso_destino, ciclo_lectivo=ciclo)
            
            if not dictados_curso.exists():
                messages.error(request, f"El curso {curso_destino} no posee materias (Dictados) asociadas en este ciclo lectivo.")
                return self.render_to_response(self.get_context_data())

            creados = 0
            errores = []

            try:
                with transaction.atomic():
                    for form in formset:
                        if not form.cleaned_data or not form.cleaned_data.get('dni'):
                            continue

                        dni = form.cleaned_data['dni'].strip()
                        apellido = form.cleaned_data['apellido'].strip()
                        nombre = form.cleaned_data['nombre'].strip()
                        
                        legajo_in = form.cleaned_data.get('numero_legajo')
                        libro_in = form.cleaned_data.get('libro')
                        folio_in = form.cleaned_data.get('folio')

                        # Validaciones preventivas de Unicidad
                        if Persona.objects.filter(dni=dni).exists():
                            errores.append(f"El DNI {dni} ya está registrado en el sistema.")
                            continue

                        # Manejo de campos requeridos por el esquema de Persona
                        legajo_final = legajo_in.strip() if legajo_in else f"LEG-{dni}-{random.randint(10,99)}"
                        if Persona.objects.filter(numero_legajo=legajo_final).exists():
                            legajo_final = f"LEG-{dni}-{random.randint(100,999)}"

                        # Fallbacks para cumplir constraints NOT NULL de la tabla gestion_academica_persona
                        cuil_dummy = f"20{dni}7" 
                        email_dummy = f"alu.{dni}@institucion.edu.ar"
                        
                        # 1. Crear el Usuario base
                        username = f"alu_{dni}"
                        user = User.objects.create_user(
                            username=username,
                            password=dni, # Contraseña por defecto = DNI
                            first_name=nombre[:30],
                            last_name=apellido[:30]
                        )

                        # 2. Insertar en gestion_academica_persona
                        persona = Persona.objects.create(
                            user=user,
                            dni=dni,
                            cuil=cuil_dummy,
                            apellido=apellido,
                            nombre=nombre,
                            numero_legajo=legajo_final,
                            email=email_dummy,
                            telefono="S/D",
                            domicilio="S/D"
                        )

                        # Fallbacks obligatorios para cumplimiento NOT NULL de gestion_academica_alumno
                        libro_final = libro_in.strip() if libro_in else "S/D"
                        folio_final = folio_in.strip() if folio_in else "S/D"

                        # 3. Insertar en gestion_academica_alumno
                        alumno = Alumno.objects.create(
                            persona=persona,
                            libro=libro_final,
                            folio=folio_final,
                            activo=True
                        )

                        # 4. Insertar en gestion_academica_historialacademico
                        HistorialAcademico.objects.create(
                            alumno=alumno,
                            curso=curso_destino,
                            ciclo_lectivo=ciclo,
                            estado_final='CURSANDO'
                        )

                        # 5. Insertar matriculaciones en gestion_academica_inscripciondictado
                        for dictado in dictados_curso:
                            InscripcionDictado.objects.create(
                                alumno=alumno,
                                dictado=dictado,
                                ciclo_lectivo=ciclo,
                                condicion='REGULAR'
                            )
                        
                        creados += 1

            except Exception as e:
                messages.error(request, f"Error de integridad en base de datos: {str(e)}")
                return self.render_to_response(self.get_context_data())

            if creados > 0:
                messages.success(request, f"Se cargaron {creados} alumnos con éxito en {curso_destino} y se les generó la inscripción automática a {dictados_curso.count()} materias.")
            if errores:
                for err in errores:
                    messages.warning(request, err)

            return redirect('/carga/alumnos/') # Poné acá el nombre de la URL o path correspondiente de tu app
        
        else:
            messages.error(request, "Error de validación en los campos del listado.")
            return self.render_to_response(self.get_context_data())