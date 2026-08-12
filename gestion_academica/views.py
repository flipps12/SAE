from django.views.generic import ListView, FormView, UpdateView, DetailView, TemplateView, CreateView
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Subquery, OuterRef, Count, Prefetch, Sum, Case, When, DecimalField, F
from django.utils.dateparse import parse_date
from django.db.models import ProtectedError
from django.db import IntegrityError

from datetime import date
import random
from decimal import Decimal, InvalidOperation

from django.forms import modelformset_factory

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook
from .forms import ImportarAlumnosForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

from .models import TIPO_ETAPA_CHOICES

from .models import Persona, Alumno, Curso, CicloLectivo, HistorialAcademico, Especialidad, InscripcionDictado, Dictado, Profesor, PersonalCargo, TipoCargo, NotaActividad, Intensificacion, NotaEtapa, Materia, Turno, Aula, Burbuja, Asistencia, Comunicado, Preceptor, AsignacionPreceptor

from .forms import AlumnoForm, PersonaForm, ImportarProfesoresForm,ProfesorUpdateForm, DictadoFormSet, HorarioFormSet, AsignacionCargoFormSet, TipoCargoForm, EspecialidadForm, CursoForm, MateriaForm, TurnoForm, AulaForm, BurbujaForm, CicloLectivoForm, AlumnoFormSet, ComunicadoForm, AltaPersonalCargoForm

User = get_user_model()


def _valor_decimal(value):
    value = str(value).strip().replace(',', '.')
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


# --------------------------------------------------------------------------------------
# ---                           ImportarAlumnosView                                  ---
# --------------------------------------------------------------------------------------

class ImportarAlumnosView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'alumnos/importar_alumnos.html'
    form_class = ImportarAlumnosForm

    def test_func(self):
        """
        Esta función define el filtro de seguridad de Django:
        Retorna True solo si el usuario está logueado y es staff (is_staff=True).
        """
        return self.request.user.is_staff
    
    def handle_no_permission(self):
        """
        Si el usuario no es staff, le muestra un mensaje de error y
        lo redirige, evitando que vea la pantalla o pueda subir archivos.
        """
        messages.error(self.request, "Acceso denegado. Solo el personal jerárquico o administrativo (Staff) puede realizar importaciones masivas.")
        return super().handle_no_permission()

    def form_valid(self, form):
        # Capturamos el archivo subido
        archivo = self.request.FILES['archivo_excel']
        resultados = None

        try:
            # Leer el archivo directo de la memoria con openpyxl
            wb = load_workbook(filename=archivo, data_only=True)
            sheet = wb.active
            
            # Mapeamos qué columna tiene cada dato según la fila 1 (Cabecera)
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1) if cell.value is not None}
            
            # Columnas obligatorias que vamos a exigir
            required_headers = [
                'username', 'password', 'nombre', 'apellido', 'dni', 'cuil', 
                'numero_legajo', 'fecha_nacimiento', 'libro', 'folio', 'activo', 
                'ciclo_lectivo', 'curso_nivel', 'curso_division', 'especialidad'
            ]
            
            # Si falta alguna columna, frenamos el proceso de inmediato
            faltantes = [req for req in required_headers if req not in headers]
            if faltantes:
                messages.error(self.request, f"Al Excel le faltan columnas críticas: {', '.join(faltantes)}")
                return self.form_invalid(form)

            exitosos = []
            errores = []

            # Recorremos desde la fila 2 en adelante
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                # Si la celda de usuario está vacía, ignoramos la fila de manera segura
                if row[headers['username'] - 1].value is None:
                    continue
                
                # Armamos un diccionario con los datos limpios de la fila actual
                data = {key: row[idx - 1].value for key, idx in headers.items()}
                alumno_str = f"{data.get('apellido', '')}, {data.get('nombre', '')} (DNI: {data.get('dni', '')})"

                try:
                    # Bloque aislado por fila: si esta falla, no afecta a los demás alumnos
                    with transaction.atomic():
                        
                        # 1. Validar infraestructura
                        ciclo = CicloLectivo.objects.get(anio=int(data['ciclo_lectivo']))
                        curso = Curso.objects.get(
                            nivel=int(data['curso_nivel']),
                            division=int(data['curso_division']),
                            especialidad__nombre__iexact=str(data['especialidad']).strip()
                        )

                        # 2. Crear credenciales de acceso
                        user, user_created = User.objects.get_or_create(
                            username=str(data['username']).strip()
                        )
                        if user_created:
                            user.set_password(str(data['password']))
                            user.save()

                        # 3. Crear datos personales base
                        persona, _ = Persona.objects.get_or_create(
                            dni=str(data['dni']).strip(),
                            defaults={
                                'user': user,
                                'nombre': str(data['nombre']).strip(),
                                'apellido': str(data['apellido']).strip(),
                                'cuil': str(data['cuil']).strip(),
                                'numero_legajo': str(data['numero_legajo']).strip(),
                                'fecha_nacimiento': data['fecha_nacimiento']
                            }
                        )

                        # 4. Crear el perfil de Alumno
                        activo_bool = True if str(data['activo']).upper() in ['TRUE', '1', 'SÍ', 'SI'] else False
                        alumno, _ = Alumno.objects.get_or_create(
                            persona=persona,
                            defaults={
                                'libro': str(data['libro'] or '').strip(),
                                'folio': str(data['folio'] or '').strip(),
                                'activo': activo_bool
                            }
                        )

                        # 5. Generar la inscripción (Historial Académico)
                        _, hist_created = HistorialAcademico.objects.get_or_create(
                            alumno=alumno,
                            ciclo_lectivo=ciclo,
                            defaults={'curso': curso, 'estado_final': 'CURSANDO'}
                        )

                        msg = f"Inscripto con éxito en {curso}" if hist_created else "El alumno ya pertenecía a este ciclo."
                        exitosos.append({'fila': row_idx, 'alumno': alumno_str, 'detalle': msg})

                except CicloLectivo.DoesNotExist:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': f"El Ciclo Lectivo {data['ciclo_lectivo']} no existe."})
                except Curso.DoesNotExist:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': f"El Curso {data['curso_nivel']}° {data['curso_division']}ª ({data['especialidad']}) no existe."})
                except Exception as e:
                    errores.append({'fila': row_idx, 'alumno': alumno_str, 'error': str(e)})

            resultados = {'exitosos': exitosos, 'errores': errores}
            messages.success(self.request, f"Procesamiento terminado. Éxitos: {len(exitosos)} | Errores: {len(errores)}")

        except Exception as e:
            messages.error(self.request, f"Error al procesar la estructura del Excel: {e}")

        # Retornamos la misma pantalla inyectando el diccionario de resultados para renderizar las alertas
        return render(self.request, self.template_name, {'form': form, 'resultados': resultados})
    

# --------------------------------------------------------------------------------------
# ---                           ImportarProfesoresView                               ---
# --------------------------------------------------------------------------------------

class ImportarProfesoresView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'profesores/importar_profesores.html'
    form_class = ImportarProfesoresForm

    def test_func(self):
        """
        Garantiza que solo usuarios Staff (personal jerárquico/administrativo)
        puedan acceder a la carga masiva de docentes.
        """
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request, 
            "Acceso denegado. Solo el personal jerárquico o administrativo (Staff) puede realizar importaciones masivas de profesores."
        )
        return super().handle_no_permission()

    def form_valid(self, form):
        archivo = self.request.FILES['archivo_excel']
        resultados = None

        try:
            # Leer el archivo desde la memoria sin persistirlo en disco
            wb = load_workbook(filename=archivo, data_only=True)
            sheet = wb.active
            
            # Mapeamos qué columna tiene cada dato según la cabecera (Fila 1)
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1) if cell.value is not None}
            
            # Columnas críticas obligatorias para el perfil del Profesor
            required_headers = [
                'username', 'password', 'nombre', 'apellido', 'dni', 'cuil', 'numero_legajo'
            ]
            
            # Si falta alguna columna obligatoria, cortamos la ejecución de inmediato
            faltantes = [req for req in required_headers if req not in headers]
            if faltantes:
                messages.error(self.request, f"Al Excel le faltan columnas críticas: {', '.join(faltantes)}")
                return self.form_invalid(form)

            exitosos = []
            errores = []

            # Recorremos los datos desde la fila 2 en adelante
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                # Si la celda clave de usuario está vacía, saltamos la línea de manera segura
                if row[headers['username'] - 1].value is None:
                    continue
                
                # Extraemos y limpiamos los valores de la fila
                data = {key: row[idx - 1].value for key, idx in headers.items()}
                profesor_str = f"{data.get('apellido', '')}, {data.get('nombre', '')} (DNI: {data.get('dni', '')})"

                try:
                    # Bloque aislado por fila: resguarda la persistencia ante fallas individuales
                    with transaction.atomic():
                        
                        # 1. Crear o recuperar credenciales de acceso (User)
                        user, user_created = User.objects.get_or_create(
                            username=str(data['username']).strip()
                        )
                        if user_created:
                            user.set_password(str(data['password']))
                            user.save()

                        # 2. Manejo de campos opcionales (Email y Teléfono)
                        email_val = str(data['email']).strip() if 'email' in headers and data.get('email') else None
                        tel_val = str(data['telefono']).strip() if 'telefono' in headers and data.get('telefono') else None

                        # 3. Crear o actualizar los datos personales base (Persona)
                        persona, persona_created = Persona.objects.get_or_create(
                            dni=str(data['dni']).strip(),
                            defaults={
                                'user': user,
                                'nombre': str(data['nombre']).strip(),
                                'apellido': str(data['apellido']).strip(),
                                'cuil': str(data['cuil']).strip(),
                                'numero_legajo': str(data['numero_legajo']).strip(),
                                'email': email_val,
                                'telefono': tel_val
                            }
                        )

                        # En caso de que la persona existiera pero sus datos cambiaron en el Excel (ej. SAE25 al 26)
                        if not persona_created:
                            persona.numero_legajo = str(data['numero_legajo']).strip()
                            if email_val: persona.email = email_val
                            if tel_val: persona.telefono = tel_val
                            persona.save()

                        # 4. Asignar el perfil de Profesor
                        profesor, prof_created = Profesor.objects.get_or_create(persona=persona)

                        msg = "Profesor dado de alta y vinculado correctamente." if prof_created else "El docente ya se encontraba registrado en el sistema."
                        exitosos.append({'fila': row_idx, 'alumno': profesor_str, 'detalle': msg})

                except Exception as e:
                    errores.append({'fila': row_idx, 'alumno': profesor_str, 'error': str(e)})

            resultados = {'exitosos': exitosos, 'errores': errores}
            messages.success(self.request, f"Procesamiento terminado. Éxitos: {len(exitosos)} | Errores: {len(errores)}")

        except Exception as e:
            messages.error(self.request, f"Error al procesar la estructura del Excel: {e}")

        # Retornamos la respuesta inyectando el diccionario de resultados tal cual lo hacés en alumnos
        return render(self.request, self.template_name, {'form': form, 'resultados': resultados})


# --------------------------------------------------------------------------------------
# ---                           ListadoAlumnosView                                   ---
# --------------------------------------------------------------------------------------

class ListadoAlumnosView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Alumno
    template_name = 'alumnos/listado_alumnos.html'
    context_object_name = 'alumnos'
    paginate_by = 25  

    def test_func(self):
        """Define la regla de acceso: usuario activo y staff."""
        return self.request.user.is_active and self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al home mostrando el mensaje de error en el div de alertas."""
        messages.error(self.request, "No tienes permisos para ver el listado de alumnos.")
        return redirect('home')

    def get_queryset(self):
        # Traemos la persona de antemano y los historiales con sus respectivos cursos
        queryset = Alumno.objects.select_related('persona').prefetch_related(
            'historiales__curso__especialidad', 
            'historiales__ciclo_lectivo'
        )

        # Captura de parámetros del buscador y los selectores
        search_query = self.request.GET.get('q', '').strip()
        curso_filter = self.request.GET.get('curso', '')
        especialidad_filter = self.request.GET.get('especialidad', '')
        estado_filter = self.request.GET.get('estado', '')

        # Filtro 1: Buscador General (Nombre, Apellido, DNI, Número de Legajo)
        if search_query:
            queryset = queryset.filter(
                Q(persona__apellido__icontains=search_query) |
                Q(persona__nombre__icontains=search_query) |
                Q(persona__dni__icontains=search_query) |
                Q(persona__numero_legajo__icontains=search_query)
            )

        # Filtro 2: Por Curso (mediante Historial Académico)
        if curso_filter:
            queryset = queryset.filter(historiales__curso_id=curso_filter)

        # Filtro 3: Por Especialidad (mediante Historial Académico)
        if cuestion_filter := especialidad_filter:
            queryset = queryset.filter(historiales__curso__especialidad_id=cuestion_filter)

        # Filtro 4: Por Estado de Actividad de la Institución
        if estado_filter == 'activos':
            queryset = queryset.filter(activo=True)
        elif estado_filter == 'inactivos':
            queryset = queryset.filter(activo=False)

        # Retornamos el queryset limpio quitando duplicados si un alumno tiene múltiples historiales
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Contexto necesario para renderizar los elementos select de los filtros
        context['cursos'] = Curso.objects.all()
        context['especialidades'] = Especialidad.objects.all()
        
        # Mantener los valores actuales en los campos del formulario tras recargar
        context['search_query'] = self.request.GET.get('q', '')
        context['curso_seleccionado'] = self.request.GET.get('curso', '')
        context['especialidad_seleccionada'] = self.request.GET.get('especialidad', '')
        context['estado_seleccionado'] = self.request.GET.get('estado', '')
        
        return context


# --------------------------------------------------------------------------------------
# ---                             EditarAlumnoView                                   ---
# --------------------------------------------------------------------------------------

class EditarAlumnoView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'alumnos/editar_alumno.html'
    success_url = reverse_lazy('listado_alumnos')

    def test_func(self):
        """Define la regla de acceso: usuario activo y staff."""
        return self.request.user.is_active and self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al home mostrando el mensaje de error."""
        messages.error(self.request, "No tienes permisos para editar este alumno.")
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()
        context['ciclo_activo'] = ciclo_activo
        
        if ciclo_activo:
            # Traemos las materias actuales ordenadas por nombre
            context['inscripciones'] = InscripcionDictado.objects.filter(
                alumno=self.object, 
                ciclo_lectivo=ciclo_activo
            ).select_related('dictado__materia', 'dictado__curso')
            
            # Traemos TODOS los dictados del año para el selector de "Agregar Materia"
            context['todos_los_dictados'] = Dictado.objects.filter(
                ciclo_lectivo=ciclo_activo
            ).select_related('materia', 'curso')

        if self.request.POST:
            context['persona_form'] = PersonaForm(self.request.POST, instance=self.object.persona)
        else:
            context['persona_form'] = PersonaForm(instance=self.object.persona)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object_with_id() # Obtenemos el alumno actual
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()

        # ACCIÓN A: Eliminar una materia suelta
        if 'action_eliminar_materia' in request.POST:
            inscripcion_id = request.POST.get('inscripcion_id')
            inscripcion = get_object_or_404(InscripcionDictado, id=inscripcion_id, alumno=self.object)
            materia_nombre = inscripcion.dictado.materia.nombre
            inscripcion.delete()
            messages.warning(request, f"Se eliminó la inscripción a la materia {materia_nombre}.")
            return render(request, self.template_name, self.get_context_data())

        # ACCIÓN B: Agregar una materia suelta (Recursante / Previa)
        if 'action_agregar_materia' in request.POST:
            dictado_id = request.POST.get('dictado_id')
            condicion = request.POST.get('condicion_materia', 'RECURSANTE')
            
            if dictado_id:
                dictado = get_object_or_404(Dictado, id=dictado_id, ciclo_lectivo=ciclo_activo)
                nueva_inscripcion, created = InscripcionDictado.objects.get_or_create(
                    alumno=self.object,
                    dictado=dictado,
                    ciclo_lectivo=ciclo_activo,
                    defaults={'condicion': condicion}
                )
                if created:
                    messages.success(request, f"Se agregó la materia {dictado.materia.nombre} como {nueva_inscripcion.get_condicion_display()}.")
                else:
                    messages.info(request, f"El alumno ya estaba inscripto en {dictado.materia.nombre}.")
            else:
                messages.error(request, "Debe seleccionar una materia válida de la lista.")
            return render(request, self.template_name, self.get_context_data())

        # ACCIÓN C: Guardado clásico de todo el formulario (Persona + Alumno)
        return super().post(request, *args, **kwargs)

    def get_object_with_id(self):
        # Auxiliar seguro para recuperar el alumno en el método post
        return get_object_or_404(Alumno, pk=self.kwargs.get(self.pk_url_kwarg or 'pk'))

    def form_valid(self, form):
        context = self.get_context_data()
        persona_form = context['persona_form']
        ciclo_activo = context['ciclo_activo']
        
        if form.is_valid() and persona_form.is_valid():
            if not ciclo_activo:
                messages.error(self.request, "No se puede guardar la inscripción porque no hay un Ciclo Lectivo activo.")
                return self.form_invalid(form)
                
            try:
                with transaction.atomic():
                    persona_form.save()
                    self.object = form.save()
                    
                    curso_seleccionado = form.cleaned_data['curso']
                    burbuja_seleccionada = form.cleaned_data['burbuja']
                    taller_seleccionado = form.cleaned_data['grupo_taller']

                    historial, _ = HistorialAcademico.objects.get_or_create(
                        alumno=self.object,
                        ciclo_lectivo=ciclo_activo,
                        defaults={'curso': curso_seleccionado}
                    )
                    historial.curso = curso_seleccionado
                    historial.burbuja = burbuja_seleccionada
                    historial.grupo_taller = taller_seleccionado
                    historial.full_clean()
                    historial.save()
                    
                    # Inscripción masiva automática (solo si no existen previamente)
                    dictados_del_curso = Dictado.objects.filter(curso=curso_seleccionado, ciclo_lectivo=ciclo_activo)
                    for dictado in dictados_del_curso:
                        InscripcionDictado.objects.get_or_create(
                            alumno=self.object,
                            dictado=dictado,
                            ciclo_lectivo=ciclo_activo,
                            defaults={'condicion': 'REGULAR'}
                        )
                    
                messages.success(self.request, "El alumno y su grilla de materias se actualizaron correctamente.")
                return super().form_valid(form)
                
            except Exception as e:
                messages.error(self.request, f"Error al guardar los datos: {str(e)}")
                return self.form_invalid(form)
        else:
            return self.form_invalid(form)


# --------------------------------------------------------------------------------------
# ---                             CrearAlumnoView                                    ---
# --------------------------------------------------------------------------------------

class CrearAlumnoView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'alumnos/crear_alumno.html'
    success_url = reverse_lazy('listado_alumnos')

    def test_func(self):
        return self.request.user.is_active and self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permisos para crear alumnos.")
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()
        context['ciclo_activo'] = ciclo_activo

        # Persona form
        context['persona_form'] = PersonaForm(
            self.request.POST or None
        )

        if ciclo_activo:
            context['todos_los_dictados'] = Dictado.objects.filter(
                ciclo_lectivo=ciclo_activo
            ).select_related('materia', 'curso')

        context['inscripciones'] = []

        return context

    def post(self, request, *args, **kwargs):
        self.object = None

        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()

        persona_form = PersonaForm(request.POST)
        alumno_form = AlumnoForm(request.POST)  # 🔥 FIX IMPORTANTE

        # UI actions (no aplican en create pero se mantienen)
        if 'action_eliminar_materia' in request.POST:
            messages.warning(request, "Primero debe crear el alumno.")
            return self.render_to_response(self.get_context_data())

        if 'action_agregar_materia' in request.POST:
            messages.info(request, "Primero debe crear el alumno.")
            return self.render_to_response(self.get_context_data())

        # VALIDACIÓN REAL
        if not persona_form.is_valid() or not alumno_form.is_valid():

            persona_errors = persona_form.errors.as_data()

            dni_exists = False
            cuil_exists = False

            # Revisamos errores reales
            if 'dni' in persona_errors:
                for e in persona_errors['dni']:
                    if 'exists' in str(e.message).lower():
                        dni_exists = True

            if 'cuil' in persona_errors:
                for e in persona_errors['cuil']:
                    if 'exists' in str(e.message).lower():
                        cuil_exists = True

            if dni_exists or cuil_exists:
                messages.error(request, "Ya existe una persona con ese DNI o CUIL.")
            else:
                messages.error(request, "Revisá los datos ingresados en el formulario.")

            return self.render_to_response(
                self.get_context_data(form=alumno_form)
            )

        if not ciclo_activo:
            messages.error(request, "No hay ciclo lectivo activo.")
            return self.render_to_response(self.get_context_data())

        try:
            with transaction.atomic():

                persona = persona_form.save()

                alumno = alumno_form.save(commit=False)
                alumno.persona = persona

                # 🔒 Forzar estado inicial
                alumno.activo = True

                alumno.save()

                curso = alumno_form.cleaned_data['curso']
                burbuja = alumno_form.cleaned_data.get('burbuja')
                grupo_taller = alumno_form.cleaned_data.get('grupo_taller')

                HistorialAcademico.objects.create(
                    alumno=alumno,
                    ciclo_lectivo=ciclo_activo,
                    curso=curso,
                    burbuja=burbuja,
                    grupo_taller=grupo_taller
                )

                dictados = Dictado.objects.filter(
                    curso=curso,
                    ciclo_lectivo=ciclo_activo
                )

                InscripcionDictado.objects.bulk_create([
                    InscripcionDictado(
                        alumno=alumno,
                        dictado=d,
                        ciclo_lectivo=ciclo_activo,
                        condicion='REGULAR'
                    )
                    for d in dictados
                ])

                if not AsignacionPreceptor.objects.filter(
                    curso=curso, ciclo_lectivo=ciclo_activo
                ).exists():
                    preceptor = (
                        Preceptor.objects
                        .exclude(asignaciones__ciclo_lectivo=ciclo_activo)
                        .first()
                    )
                    if not preceptor:
                        preceptor = (
                            Preceptor.objects
                            .annotate(n=Count('asignaciones'))
                            .order_by('n', 'id')
                            .first()
                        )
                    if preceptor:
                        AsignacionPreceptor.objects.create(
                            preceptor=preceptor,
                            curso=curso,
                            ciclo_lectivo=ciclo_activo,
                        )

            messages.success(request, "Alumno creado correctamente.")
            return redirect(self.success_url)

        except Exception as e:
            messages.error(request, f"Error al crear alumno: {e}")
            return self.render_to_response(self.get_context_data())

    def form_valid(self, form):
        # no se usa
        return super().form_valid(form)

# --------------------------------------------------------------------------------------
# ---                             ProfesorListView                                   ---
# --------------------------------------------------------------------------------------

class ProfesorListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Profesor
    template_name = 'profesores/listado_profesor.html'
    context_object_name = 'profesores'
    paginate_by = 25

    def test_func(self):
        """Solo los usuarios staff pueden ver este listado."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige si el usuario no tiene permisos."""
        messages.error(self.request, "No tienes permisos para ver el listado de profesores.")
        return redirect('home')  # Ajusta a tu URL de inicio o dashboard

    def get_queryset(self):
        # 1. Obtenemos el ciclo lectivo institucional activo
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()
        
        # 2. Creamos una subconsulta para contar de forma aislada
        subconsulta_conteo = Dictado.objects.filter(profesor=OuterRef('pk'))
        if ciclo_activo:
            subconsulta_conteo = subconsulta_conteo.filter(ciclo_lectivo=ciclo_activo)
            
        # 3. Anotamos el conteo limpio al queryset principal
        queryset = Profesor.objects.select_related('persona').annotate(
            total_materias=Subquery(
                subconsulta_conteo.values('profesor')
                .annotate(count=Count('id'))
                .values('count')
            )
        )
        
        # --- Parámetros de filtros ---
        query = self.request.GET.get('q')
        curso_id = self.request.GET.get('curso')
        grupo_ciclo = self.request.GET.get('grupo_ciclo')
        especialidad_id = self.request.GET.get('especialidad')
        
        if query:
            queryset = queryset.filter(
                Q(persona__nombre__icontains=query) |
                Q(persona__apellido__icontains=query) |
                Q(persona__dni__icontains=query) |
                Q(persona__numero_legajo__icontains=query)
            )
            
        if ciclo_activo:
            if curso_id:
                queryset = queryset.filter(dictados__curso_id=curso_id, dictados__ciclo_lectivo=ciclo_activo)
                
            if grupo_ciclo == 'basico':
                queryset = queryset.filter(dictados__curso__nivel__lte=3, dictados__ciclo_lectivo=ciclo_activo)
            elif grupo_ciclo == 'superior':
                queryset = queryset.filter(dictados__curso__nivel__gte=4, dictados__ciclo_lectivo=ciclo_activo)
                
            if especialidad_id:
                if especialidad_id == 'troncal':
                    queryset = queryset.filter(dictados__materia__especialidad__isnull=True, dictados__ciclo_lectivo=ciclo_activo)
                else:
                    queryset = queryset.filter(dictados__materia__especialidad_id=especialidad_id, dictados__ciclo_lectivo=ciclo_activo)

        return queryset.distinct().order_by('persona__apellido', 'persona__nombre', 'id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos'] = Curso.objects.all()
        context['especialidades'] = Especialidad.objects.all()
        context['search_query'] = self.request.GET.get('q', '')
        context['curso_seleccionado'] = self.request.GET.get('curso', '')
        context['grupo_ciclo_seleccionado'] = self.request.GET.get('grupo_ciclo', '')
        context['especialidad_seleccionada'] = self.request.GET.get('especialidad', '')
        return context
    
# --------------------------------------------------------------------------------------
# ---                             ListaCargosView                                   ---
# --------------------------------------------------------------------------------------

class ListaCargosView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = PersonalCargo
    template_name = 'cargos/lista_cargos.html'
    context_object_name = 'personal_con_cargos'
    paginate_by = 10 

    def test_func(self):
        """Define la regla de acceso: usuario activo y staff."""
        return self.request.user.is_active and self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al home mostrando el mensaje de error."""
        messages.error(self.request, "No tienes permisos para ver el listado de cargos.")
        return redirect('home')

    def get_queryset(self):
        # Base optimizada de la consulta
        queryset = PersonalCargo.objects.select_related('persona').prefetch_related(
            'asignaciones__cargo'
        ).filter(asignaciones__activo=True).distinct()

        # Capturamos el buscador general 'q'
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(persona__nombre__icontains=query) |
                Q(persona__apellido__icontains=query) |
                Q(persona__cuil__icontains=query) |
                Q(persona__numero_legajo__icontains=query)
            )
        
        # Capturamos el filtro por cargo
        cargo_id = self.request.GET.get('cargo')
        if cargo_id:
            queryset = queryset.filter(asignaciones__cargo_id=cargo_id)
            
        return queryset.order_by('persona__apellido', 'persona__nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Mantenemos los valores en el formulario tras recargar/filtrar
        context['search_query'] = self.request.GET.get('q', '')
        context['cargo_seleccionado'] = self.request.GET.get('cargo', '')
        
        # Enviamos la lista de cargos para poblar el dropdown
        context['cargos'] = TipoCargo.objects.all().order_by('nombre')
        
        return context
    
# --------------------------------------------------------------------------------------
# ---                                VerAlumnoView                                   ---
# --------------------------------------------------------------------------------------

class VerAlumnoView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Alumno
    template_name = 'alumnos/ver_alumno.html'
    context_object_name = 'alumno'

    def test_func(self):
        """Define la regla de acceso: usuario activo y staff."""
        return self.request.user.is_active and self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al home mostrando el mensaje de error."""
        messages.error(self.request, "No tienes permisos para ver el detalle de este alumno.")
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ciclo_activo = CicloLectivo.objects.filter(activo=True).first()
        context['ciclo_activo'] = ciclo_activo
        
        if ciclo_activo:
            # Buscamos el historial del alumno para este ciclo
            context['historial'] = HistorialAcademico.objects.filter(
                alumno=self.object, 
                ciclo_lectivo=ciclo_activo
            ).first()
            
            # Inscripciones a materias
            context['inscripciones'] = InscripcionDictado.objects.filter(
                alumno=self.object, 
                ciclo_lectivo=ciclo_activo
            ).select_related('dictado__materia', 'dictado__curso')
            
        return context
    
# --------------------------------------------------------------------------------------
# ---                                EditarProfesorView                              ---
# --------------------------------------------------------------------------------------

class EditarProfesorView(UserPassesTestMixin, UpdateView):
    model = Profesor
    template_name = 'profesores/editar_profesor.html'
    form_class = ProfesorUpdateForm
    success_url = reverse_lazy('listado_profesor')

    def test_func(self):
        """Verifica que el usuario sea staff para acceder a la vista."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige si el usuario no tiene permisos."""
        messages.error(self.request, "No tienes permisos de administrador para realizar esta acción.")
        return redirect('listado_profesor')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['persona_form'] = PersonaForm(self.request.POST, instance=self.object.persona)
            context['formset'] = DictadoFormSet(self.request.POST, instance=self.object)
        else:
            context['persona_form'] = PersonaForm(instance=self.object.persona)
            context['formset'] = DictadoFormSet(instance=self.object)

        # Preparar formsets de horarios para cada formulario del dictado
        for form in context['formset']:
            prefix = f"dictado_{form.instance.pk if form.instance.pk else form.prefix}"
            form.horario_formset = HorarioFormSet(
                self.request.POST or None, 
                instance=form.instance, 
                prefix=prefix
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        persona_form = context['persona_form']
        dictado_formset = context['formset']

        if persona_form.is_valid() and form.is_valid() and dictado_formset.is_valid():
            with transaction.atomic():
                persona_form.save()
                form.save()
                
                # Guardar dictados
                dictados = dictado_formset.save(commit=False)
                for instance in dictados:
                    instance.save()
                dictado_formset.save_m2m()

                # Guardar horarios anidados
                for dictado_form in dictado_formset.forms:
                    prefix = f"dictado_{dictado_form.instance.pk if dictado_form.instance.pk else dictado_form.prefix}"
                    h_formset = HorarioFormSet(self.request.POST, instance=dictado_form.instance, prefix=prefix)
                    
                    if h_formset.is_valid():
                        h_formset.save()
                    else:
                        messages.error(self.request, f"Error en horarios de {dictado_form.instance}: {h_formset.errors}")
                        return self.render_to_response(context)
            
            messages.success(self.request, "Datos guardados correctamente.")
            return redirect(self.success_url)
        
        return self.render_to_response(context)


# --------------------------------------------------------------------------------------
# ---                                VerProfesorView                                 ---
# --------------------------------------------------------------------------------------

class DetalleProfesorView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Profesor
    template_name = 'profesores/ver_profesor.html'
    context_object_name = 'profesor'

    def test_func(self):
        """Si solo quieres que el staff vea esto, mantenlo."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al home mostrando el mensaje de error."""
        messages.error(self.request, "No tienes permisos para ver el detalle de este profesor.")
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Accedemos a los dictados del profesor mediante la relación definida
        # Usamos prefetch_related para optimizar la consulta de horarios
        context['dictados'] = self.object.dictados.all().prefetch_related('horarios')
        return context
    



# --------------------------------------------------------------------------------------
# ---                              CalificacionesView                                ---
# --------------------------------------------------------------------------------------

class ListaCalificacionesView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = InscripcionDictado
    template_name = 'calificaciones/lista_calificaciones.html'
    context_object_name = 'inscripciones'

    def test_func(self):
        """Solo el personal staff puede acceder."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Evita el 403 y muestra un mensaje amigable."""
        messages.error(
            self.request,
            "No tienes permisos para acceder al listado de calificaciones."
        )
        return redirect('home')

    def get_queryset(self):
        # 1. Filtros base con carga optimizada
        queryset = InscripcionDictado.objects.select_related(
            'alumno__persona',
            'dictado__materia',
            'dictado__curso',
            'ciclo_lectivo'
        ).prefetch_related(
            'dictado__notas_etapas',
            'dictado__notas_actividades',
            'dictado__intensificaciones'
        )

        # 2. Captura de parámetros GET
        ciclo_id = self.request.GET.get('ciclo')
        curso_id = self.request.GET.get('curso')
        dictado_id = self.request.GET.get('dictado')
        search_query = self.request.GET.get('q')

        # 3. Aplicación de filtros
        if ciclo_id:
            queryset = queryset.filter(ciclo_lectivo_id=ciclo_id)
        else:
            queryset = queryset.filter(ciclo_lectivo__activo=True)

        if curso_id:
            queryset = queryset.filter(dictado__curso_id=curso_id)

        if dictado_id:
            queryset = queryset.filter(dictado_id=dictado_id)

        if search_query:
            queryset = queryset.filter(
                Q(alumno__persona__nombre__icontains=search_query) |
                Q(alumno__persona__apellido__icontains=search_query)
            )

        # 4. Exclusión de filas sin registros (notas)
        queryset = queryset.filter(
            Q(dictado__notas_etapas__isnull=False) |
            Q(dictado__notas_actividades__isnull=False) |
            Q(dictado__intensificaciones__isnull=False)
        ).distinct().order_by(
            'dictado__curso__nivel',
            'dictado__curso__division',
            'alumno__persona__apellido'
        )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ciclo_id = self.request.GET.get('ciclo')
        curso_id = self.request.GET.get('curso')

        dictados_queryset = Dictado.objects.all()

        if ciclo_id:
            dictados_queryset = dictados_queryset.filter(ciclo_lectivo_id=ciclo_id)
        else:
            dictados_queryset = dictados_queryset.filter(ciclo_lectivo__activo=True)

        if curso_id:
            dictados_queryset = dictados_queryset.filter(curso_id=curso_id)

        context['ciclos'] = CicloLectivo.objects.all().order_by('-anio')
        context['cursos'] = Curso.objects.all()
        context['dictados'] = dictados_queryset.order_by('materia__nombre')

        context['selected_ciclo'] = (
            int(ciclo_id)
            if ciclo_id and ciclo_id.isdigit()
            else None
        )

        context['selected_curso'] = (
            int(curso_id)
            if curso_id and curso_id.isdigit()
            else None
        )

        context['selected_dictado'] = (
            int(self.request.GET.get('dictado'))
            if self.request.GET.get('dictado')
            and self.request.GET.get('dictado').isdigit()
            else None
        )

        context['search_query'] = self.request.GET.get('q', '')

        return context
    
# --------------------------------------------------------------------------------------
# ---                   DetalleCalificacionesAlumnoView                              ---
# --------------------------------------------------------------------------------------


class DetalleCalificacionesAlumnoView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Alumno
    template_name = 'calificaciones/detalle_alumno.html'
    context_object_name = 'alumno'

    def test_func(self):
        """Solo el personal staff puede ver el detalle de calificaciones."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al inicio mostrando el mensaje de error."""
        messages.error(
            self.request,
            "No tienes permisos para ver el detalle de las calificaciones."
        )
        return redirect('home')

    def get_queryset(self):
        # Traemos la persona de un viaje y preparamos el prefetch de sus materias cursadas
        return Alumno.objects.select_related('persona').prefetch_related(
            Prefetch(
                'inscripciones_dictados',
                queryset=InscripcionDictado.objects.select_related(
                    'dictado__materia',
                    'dictado__curso__especialidad',
                    'ciclo_lectivo'
                ).prefetch_related(
                    'dictado__notas_actividades',
                    'dictado__notas_etapas',
                    'dictado__intensificaciones'
                ).order_by(
                    '-ciclo_lectivo__anio',
                    'dictado__materia__nombre'
                ),
                to_attr='historial_materias'
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        materias_por_ciclo = {}

        for inscripcion in self.object.historial_materias:
            anio = inscripcion.ciclo_lectivo.anio

            if anio not in materias_por_ciclo:
                materias_por_ciclo[anio] = {
                    'curso': inscripcion.dictado.curso,
                    'inscripciones': []
                }

            materias_por_ciclo[anio]['inscripciones'].append(inscripcion)

        context['historial_agrupado'] = materias_por_ciclo
        return context
    

# --------------------------------------------------------------------------------------
# ---                           EditarPersonalView                                   ---
# --------------------------------------------------------------------------------------

class EditarPersonalView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'cargos/editar_personal.html'

    def test_func(self):
        """Solo usuarios staff pueden editar."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige mostrando el mensaje en lugar del 403."""
        messages.error(
            self.request,
            "No tienes permisos para editar el personal."
        )
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        persona = get_object_or_404(Persona, pk=self.kwargs['pk'])
        perfil_cargo, _ = PersonalCargo.objects.get_or_create(persona=persona)

        context['persona'] = persona

        # Si venimos de un POST inválido reutilizamos los formularios
        context['form_persona'] = kwargs.get(
            'form_persona',
            PersonaForm(instance=persona)
        )

        context['formset'] = kwargs.get(
            'formset',
            AsignacionCargoFormSet(instance=perfil_cargo)
        )

        return context

    def post(self, request, *args, **kwargs):
        persona = get_object_or_404(Persona, pk=self.kwargs['pk'])
        perfil_cargo, _ = PersonalCargo.objects.get_or_create(persona=persona)

        form_persona = PersonaForm(request.POST, instance=persona)
        formset = AsignacionCargoFormSet(request.POST, instance=perfil_cargo)

        if form_persona.is_valid() and formset.is_valid():
            form_persona.save()
            formset.save()

            messages.success(request, "Datos guardados correctamente.")
            return redirect('lista_cargos')

        print("Errores form_persona:", form_persona.errors)
        print("Errores formset:", formset.errors)

        return self.render_to_response(
            self.get_context_data(
                form_persona=form_persona,
                formset=formset,
                **kwargs
            )
        )
    

# --------------------------------------------------------------------------------------
# ---                           DetallePersonalCargoView                             ---
# --------------------------------------------------------------------------------------

class DetallePersonalCargoView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Persona
    template_name = 'cargos/ver_personal_cargos.html'
    context_object_name = 'persona'

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos para ver el detalle del personal."
        )
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 1. Recuperar el filtro de cargo si existe
        cargo_id = self.request.GET.get('cargo')

        # 2. Recuperar el perfil de cargo de la persona
        perfil = getattr(self.object, 'perfil_cargo', None)

        # 3. Filtrar las asignaciones
        if perfil:
            asignaciones = perfil.asignaciones.all()
            if cargo_id:
                asignaciones = asignaciones.filter(cargo_id=cargo_id)
            context['asignaciones'] = asignaciones
        else:
            context['asignaciones'] = []

        # 4. Preparar la lista de cargos para el filtro
        cargos = TipoCargo.objects.all()
        for c in cargos:
            c.es_seleccionado = (str(c.id) == str(cargo_id))

        context['cargos'] = cargos

        return context
    
# --------------------------------------------------------------------------------------
# ---                           EditarCalificacionesView                             ---
# --------------------------------------------------------------------------------------

class EditarCalificacionesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "calificaciones/editar_calificaciones.html"

    def test_func(self):
        """Solo el personal staff puede editar calificaciones."""
        return self.request.user.is_staff

    def handle_no_permission(self):
        """Redirige al inicio mostrando el mensaje de error."""
        messages.error(
            self.request,
            "No tienes permisos para editar calificaciones."
        )
        return redirect("home")

    def get_formsets(self, dictado, alumno, data=None):

        EtapaFS = modelformset_factory(
            NotaEtapa,
            fields=("etapa", "valor_numerico", "valor_conceptual", "valoracion_previa"),
            extra=0,
            can_delete=False
        )

        ActividadFS = modelformset_factory(
            NotaActividad,
            fields=("nombre_actividad", "valor", "cuatrimestre", "fecha"),
            extra=1,
            can_delete=True
        )

        IntensifFS = modelformset_factory(
            Intensificacion,
            fields=("fecha", "valor"),
            extra=1,
            can_delete=True
        )

        # asegurar etapas
        for etapa, _ in TIPO_ETAPA_CHOICES:
            NotaEtapa.objects.get_or_create(
                dictado=dictado,
                alumno=alumno,
                etapa=etapa,
                defaults={"valor_numerico": 0}
            )

        return {
            "etapa": EtapaFS(
                data,
                queryset=NotaEtapa.objects.filter(dictado=dictado, alumno=alumno),
                prefix="etapa"
            ),
            "act": ActividadFS(
                data,
                queryset=NotaActividad.objects.filter(dictado=dictado, alumno=alumno),
                prefix="act"
            ),
            "int": IntensifFS(
                data,
                queryset=Intensificacion.objects.filter(dictado=dictado, alumno=alumno),
                prefix="int"
            )
        }

    def get(self, request, *args, **kwargs):
        dictado = get_object_or_404(Dictado, pk=kwargs["dictado_id"])
        alumno = get_object_or_404(Alumno, pk=kwargs["alumno_id"])

        formsets = self.get_formsets(dictado, alumno)

        context = self.get_context_data(**kwargs)
        context.update({
            "formsets": formsets,
            "alumno": alumno,
            "dictado": dictado
        })

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        dictado = get_object_or_404(Dictado, pk=kwargs["dictado_id"])
        alumno = get_object_or_404(Alumno, pk=kwargs["alumno_id"])

        formsets = self.get_formsets(dictado, alumno, request.POST)

        valid = all(fs.is_valid() for fs in formsets.values())

        if valid:
            formsets["etapa"].save()

            act_forms = formsets["act"].save(commit=False)
            for obj in act_forms:
                obj.dictado = dictado
                obj.alumno = alumno
                obj.save()
            for obj in formsets["act"].deleted_objects:
                obj.delete()

            int_forms = formsets["int"].save(commit=False)
            for obj in int_forms:
                obj.dictado = dictado
                obj.alumno = alumno
                obj.save()
            for obj in formsets["int"].deleted_objects:
                obj.delete()

            messages.success(request, "Guardado OK")
            return redirect("lista_calificaciones")

        context = self.get_context_data(**kwargs)
        context.update({
            "formsets": formsets,
            "alumno": alumno,
            "dictado": dictado
        })

        return self.render_to_response(context)
    

# --------------------------------------------------------------------------------------
# ---                              CrearProfesorView                                 ---
# --------------------------------------------------------------------------------------

class CrearProfesorView(UserPassesTestMixin, CreateView):
    model = Profesor
    form_class = ProfesorUpdateForm
    template_name = "profesores/crear_profesor.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("listado_profesor")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.request.POST:
            context["persona_form"] = PersonaForm(self.request.POST)
        else:
            context["persona_form"] = PersonaForm()

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        persona_form = context["persona_form"]

        if persona_form.is_valid() and form.is_valid():

            with transaction.atomic():

                persona = persona_form.save()

                profesor = form.save(commit=False)
                profesor.persona = persona
                profesor.save()

            messages.success(
                self.request,
                "Profesor creado correctamente."
            )

            accion = self.request.POST.get("accion")

            if accion == "nuevo":
                return redirect("crear_profesor")

            return redirect("editar_profesor", pk=profesor.pk)

        return self.render_to_response(context)

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))
    

# --------------------------------------------------------------------------------------
# ---                              GestionTipoCargoView                              ---
# --------------------------------------------------------------------------------------

class GestionTipoCargoView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_tipo_cargo.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            cargo = get_object_or_404(TipoCargo, pk=pk)
            context["form"] = TipoCargoForm(instance=cargo)
            context["editando"] = cargo
        else:
            context["form"] = TipoCargoForm()

        context["tipos"] = TipoCargo.objects.order_by("nombre")

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            cargo = get_object_or_404(TipoCargo, pk=pk)
            form = TipoCargoForm(request.POST, instance=cargo)
        else:
            form = TipoCargoForm()

        if request.POST.get("accion") == "eliminar":

            cargo = get_object_or_404(TipoCargo, pk=pk)

            try:
                cargo.delete()
                messages.success(request, "Cargo eliminado correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No puede eliminarse porque está siendo utilizado."
                )

            return redirect("gestion_tipo_cargo")

        form = TipoCargoForm(request.POST, instance=cargo if pk else None)

        if form.is_valid():
            form.save()

            if pk:
                messages.success(request, "Cargo actualizado correctamente.")
            else:
                messages.success(request, "Cargo creado correctamente.")

            return redirect("gestion_tipo_cargo")

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)
    
# --------------------------------------------------------------------------------------
# ---                            GestionEspecialidadView                             ---
# --------------------------------------------------------------------------------------

class GestionEspecialidadView(UserPassesTestMixin, TemplateView):
        template_name = "cruds/gestion_especialidad.html"

        def test_func(self):
            return self.request.user.is_staff

        def handle_no_permission(self):
            messages.error(
                self.request,
                "No tienes permisos de administrador para realizar esta acción."
            )
            return redirect("home")

        def get_context_data(self, **kwargs):
            context = super().get_context_data(**kwargs)

            pk = self.request.GET.get("editar")

            if pk:
                especialidad = get_object_or_404(Especialidad, pk=pk)
                context["form"] = EspecialidadForm(instance=especialidad)
                context["editando"] = especialidad
            else:
                context["form"] = EspecialidadForm()

            context["especialidades"] = Especialidad.objects.order_by("nombre")

            return context

        def post(self, request, *args, **kwargs):

            pk = request.POST.get("pk")

            if pk:
                especialidad = get_object_or_404(Especialidad, pk=pk)
                form = EspecialidadForm(request.POST, instance=especialidad)
            else:
                especialidad = None
                form = EspecialidadForm(request.POST)

            if request.POST.get("accion") == "eliminar":

                especialidad = get_object_or_404(Especialidad, pk=pk)

                try:
                    especialidad.delete()
                    messages.success(request, "Especialidad eliminada correctamente.")
                except ProtectedError:
                    messages.error(
                        request,
                        "No puede eliminarse porque está siendo utilizada."
                    )

                return redirect("gestion_especialidad")

            if form.is_valid():
                form.save()

                if pk:
                    messages.success(request, "Especialidad actualizada correctamente.")
                else:
                    messages.success(request, "Especialidad creada correctamente.")

                return redirect("gestion_especialidad")

            context = self.get_context_data()
            context["form"] = form

            return self.render_to_response(context)
        
# --------------------------------------------------------------------------------------
# ---                              GestionCursoView                                  ---
# --------------------------------------------------------------------------------------

class GestionCursoView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_curso.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            curso = get_object_or_404(Curso, pk=pk)
            context["form"] = CursoForm(instance=curso)
            context["editando"] = curso
        else:
            context["form"] = CursoForm()

        context["cursos"] = (
            Curso.objects
            .select_related("especialidad")
            .order_by("nivel", "division", "especialidad__nombre")
        )

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            curso = get_object_or_404(Curso, pk=pk)
            form = CursoForm(request.POST, instance=curso)
        else:
            curso = None
            form = CursoForm(request.POST)

        if request.POST.get("accion") == "eliminar":

            curso = get_object_or_404(Curso, pk=pk)

            try:
                curso.delete()
                messages.success(request, "Curso eliminado correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No puede eliminarse porque existen alumnos inscriptos o historial académico asociado."
                )

            return redirect("gestion_curso")

        if form.is_valid():

            form.save()

            if pk:
                messages.success(request, "Curso actualizado correctamente.")
            else:
                messages.success(request, "Curso creado correctamente.")

            return redirect("gestion_curso")

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)
    
# --------------------------------------------------------------------------------------
# ---                            GestionMateriaView                                  ---
# --------------------------------------------------------------------------------------

class GestionMateriaView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_materia.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            materia = get_object_or_404(Materia, pk=pk)
            context["form"] = MateriaForm(instance=materia)
            context["editando"] = materia
        else:
            context["form"] = MateriaForm()

        context["materias"] = (
            Materia.objects
            .select_related("especialidad")
            .order_by("nombre", "especialidad__nombre")
        )

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            materia = get_object_or_404(Materia, pk=pk)
            form = MateriaForm(request.POST, instance=materia)
        else:
            materia = None
            form = MateriaForm(request.POST)

        # ===== ELIMINAR =====
        if request.POST.get("accion") == "eliminar":
            materia = get_object_or_404(Materia, pk=pk)

            try:
                materia.delete()
                messages.success(request, "Materia eliminada correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No puede eliminarse porque está siendo utilizada en otros registros."
                )

            return redirect("gestion_materia")

        # ===== GUARDAR / EDITAR =====
        if form.is_valid():
            form.save()

            if pk:
                messages.success(request, "Materia actualizada correctamente.")
            else:
                messages.success(request, "Materia creada correctamente.")

            return redirect("gestion_materia")

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)
    
# --------------------------------------------------------------------------------------
# ---                              GestionTurnoView                                  ---
# --------------------------------------------------------------------------------------

class GestionTurnoView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_turno.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            turno = get_object_or_404(Turno, pk=pk)
            context["form"] = TurnoForm(instance=turno)
            context["editando"] = turno
        else:
            context["form"] = TurnoForm()

        context["turnos"] = Turno.objects.order_by("nombre")

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            turno = get_object_or_404(Turno, pk=pk)
            form = TurnoForm(request.POST, instance=turno)
        else:
            turno = None
            form = TurnoForm(request.POST)

        # ===== ELIMINAR =====
        if request.POST.get("accion") == "eliminar":
            turno = get_object_or_404(Turno, pk=pk)

            try:
                turno.delete()
                messages.success(request, "Turno eliminado correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No puede eliminarse porque está siendo utilizado."
                )

            return redirect("gestion_turno")

        # ===== GUARDAR =====
        if form.is_valid():
            form.save()

            if pk:
                messages.success(request, "Turno actualizado correctamente.")
            else:
                messages.success(request, "Turno creado correctamente.")

            return redirect("gestion_turno")

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)
    
# --------------------------------------------------------------------------------------
# ---                              GestionAulaView                                   ---
# --------------------------------------------------------------------------------------


class GestionAulaView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_aula.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            aula = get_object_or_404(Aula, pk=pk)
            context["form"] = AulaForm(instance=aula)
            context["editando"] = aula
        else:
            context["form"] = AulaForm()

        context["aulas"] = Aula.objects.order_by("nombre")

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            aula = get_object_or_404(Aula, pk=pk)
            form = AulaForm(request.POST, instance=aula)
        else:
            form = AulaForm(request.POST)

        if request.POST.get("accion") == "eliminar":
            aula = get_object_or_404(Aula, pk=pk)

            try:
                aula.delete()
                messages.success(request, "Aula eliminada correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No se puede eliminar porque está siendo utilizada."
                )

            return redirect("gestion_aula")

        if form.is_valid():
            form.save()

            if pk:
                messages.success(request, "Aula actualizada correctamente.")
            else:
                messages.success(request, "Aula creada correctamente.")

            return redirect("gestion_aula")

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)

# --------------------------------------------------------------------------------------
# ---                             GestionBurbujaView                                 ---
# --------------------------------------------------------------------------------------   


class GestionBurbujaView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_burbuja.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            burbuja = get_object_or_404(Burbuja, pk=pk)
            context["form"] = BurbujaForm(instance=burbuja)
            context["editando"] = burbuja
        else:
            context["form"] = BurbujaForm()

        context["burbujas"] = (
            Burbuja.objects
            .select_related("ciclo_lectivo")
            .order_by("ciclo_lectivo__anio", "nombre")
        )

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            burbuja = get_object_or_404(Burbuja, pk=pk)
            form = BurbujaForm(request.POST, instance=burbuja)
        else:
            form = BurbujaForm(request.POST)

        # ===== ELIMINAR =====
        if request.POST.get("accion") == "eliminar":
            burbuja = get_object_or_404(Burbuja, pk=pk)

            try:
                burbuja.delete()
                messages.success(request, "Burbuja eliminada correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No se puede eliminar porque está siendo utilizada."
                )

            return redirect("gestion_burbuja")

        # ===== GUARDAR =====
        try:
            if form.is_valid():
                form.save()

                if pk:
                    messages.success(request, "Burbuja actualizada correctamente.")
                else:
                    messages.success(request, "Burbuja creada correctamente.")

                return redirect("gestion_burbuja")

        except IntegrityError:
            messages.error(
                request,
                "Ya existe una burbuja con ese nombre en este ciclo lectivo."
            )

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)

# --------------------------------------------------------------------------------------
# ---                           GestionCicloLectivoView                              ---
# --------------------------------------------------------------------------------------  

class GestionCicloLectivoView(UserPassesTestMixin, TemplateView):
    template_name = "cruds/gestion_ciclo_lectivo.html"

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(
            self.request,
            "No tienes permisos de administrador para realizar esta acción."
        )
        return redirect("home")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.request.GET.get("editar")

        if pk:
            ciclo = get_object_or_404(CicloLectivo, pk=pk)
            context["form"] = CicloLectivoForm(instance=ciclo)
            context["editando"] = ciclo
        else:
            context["form"] = CicloLectivoForm()

        context["ciclos"] = CicloLectivo.objects.order_by("-anio")

        return context

    def post(self, request, *args, **kwargs):

        pk = request.POST.get("pk")

        if pk:
            ciclo = get_object_or_404(CicloLectivo, pk=pk)
            form = CicloLectivoForm(request.POST, instance=ciclo)
        else:
            form = CicloLectivoForm(request.POST)

        # ===== ELIMINAR =====
        if request.POST.get("accion") == "eliminar":
            ciclo = get_object_or_404(CicloLectivo, pk=pk)

            try:
                ciclo.delete()
                messages.success(request, "Ciclo lectivo eliminado correctamente.")
            except ProtectedError:
                messages.error(
                    request,
                    "No se puede eliminar porque está siendo utilizado."
                )

            return redirect("gestion_ciclo_lectivo")

        # ===== GUARDAR =====
        try:
            if form.is_valid():
                form.save()

                if pk:
                    messages.success(request, "Ciclo lectivo actualizado correctamente.")
                else:
                    messages.success(request, "Ciclo lectivo creado correctamente.")

                return redirect("gestion_ciclo_lectivo")

        except IntegrityError:
            messages.error(
                request,
                "Ya existe un ciclo lectivo activo o el año ya está registrado."
            )

        context = self.get_context_data()
        context["form"] = form

        return self.render_to_response(context)
    

# --------------------------------------------------------------------------------------
# ---                           MisDictadosProfesorView                              ---
# -------------------------------------------------------------------------------------- 

class MisDictadosProfesorView(LoginRequiredMixin, ListView):
    model = Dictado
    template_name = "profesores/mis_dictados.html"
    context_object_name = "dictados"
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user

        # Obtener profesor asociado
        try:
            profesor = user.persona.perfil_profesor
        except:
            return Dictado.objects.none()

        queryset = (
            Dictado.objects
            .select_related(
                'materia',
                'curso',
                'curso__especialidad',
                'ciclo_lectivo',
                'profesor'
            )
            .filter(profesor=profesor)
            .order_by('curso__nivel', 'curso__division', 'materia__nombre')
        )

        # -----------------------------
        # FILTROS
        # -----------------------------
        curso = self.request.GET.get("curso")
        especialidad = self.request.GET.get("especialidad")
        search = self.request.GET.get("q")

        if curso:
            queryset = queryset.filter(curso_id=curso)

        if especialidad:
            queryset = queryset.filter(curso__especialidad_id=especialidad)

        if search:
            queryset = queryset.filter(
                Q(materia__nombre__icontains=search) |
                Q(curso__nivel__icontains=search) |
                Q(curso__division__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user

        try:
            profesor = user.persona.perfil_profesor
        except:
            profesor = None

        if profesor:
            dictados = (
                Dictado.objects
                .filter(profesor=profesor)
                .select_related("curso", "curso__especialidad")
            )

            context["cursos"] = (
                Curso.objects
                .filter(dictados__profesor=profesor)
                .distinct()
                .order_by("nivel", "division")
            )

            context["especialidades"] = (
                Especialidad.objects
                .filter(curso__dictados__profesor=profesor)
                .distinct()
                .order_by("nombre")
            )
        else:
            context["cursos"] = Curso.objects.none()
            context["especialidades"] = Especialidad.objects.none()

        context["curso_selected"] = self.request.GET.get("curso", "")
        context["especialidad_selected"] = self.request.GET.get("especialidad", "")
        context["search"] = self.request.GET.get("q", "")

        return context

# --------------------------------------------------------------------------------------
# ---                             DictadoAlumnosView                                 ---
# -------------------------------------------------------------------------------------- 

class DictadoAlumnosView(LoginRequiredMixin, DetailView):
    model = Dictado
    template_name = "profesores/dictado_alumnos.html"
    context_object_name = "dictado"

    def get_queryset(self):
        profesor = self.request.user.persona.perfil_profesor

        return (
            Dictado.objects
            .select_related(
                "materia",
                "curso",
                "curso__especialidad",
                "ciclo_lectivo",
            )
            .filter(profesor=profesor)
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        alumnos = (
            InscripcionDictado.objects
            .filter(dictado=self.object)
            .select_related(
                "alumno",
                "alumno__persona",
            )
            .order_by(
                "alumno__persona__apellido",
                "alumno__persona__nombre"
            )
        )

        context["inscripciones"] = alumnos

        return context

# --------------------------------------------------------------------------------------
# ---                        ProfesorEditarCalificacionesView                        ---
# -------------------------------------------------------------------------------------- 

class ProfesorEditarCalificacionesView(LoginRequiredMixin, TemplateView):
    template_name = "profesores/editar_calificaciones.html"

    def get_formsets(self, dictado, alumno, data=None):

        EtapaFS = modelformset_factory(
            NotaEtapa,
            fields=("etapa", "valor_numerico", "valor_conceptual", "valoracion_previa"),
            extra=0,
            can_delete=False
        )

        ActividadFS = modelformset_factory(
            NotaActividad,
            fields=("nombre_actividad", "valor", "cuatrimestre", "fecha"),
            extra=1,
            can_delete=True
        )

        IntensifFS = modelformset_factory(
            Intensificacion,
            fields=("fecha", "valor"),
            extra=1,
            can_delete=True
        )

        # Asegurar que existan todas las etapas
        for etapa, _ in TIPO_ETAPA_CHOICES:
            NotaEtapa.objects.get_or_create(
                dictado=dictado,
                alumno=alumno,
                etapa=etapa,
                defaults={"valor_numerico": 0}
            )

        return {
            "etapa": EtapaFS(
                data,
                queryset=NotaEtapa.objects.filter(
                    dictado=dictado,
                    alumno=alumno
                ),
                prefix="etapa"
            ),
            "act": ActividadFS(
                data,
                queryset=NotaActividad.objects.filter(
                    dictado=dictado,
                    alumno=alumno
                ),
                prefix="act"
            ),
            "int": IntensifFS(
                data,
                queryset=Intensificacion.objects.filter(
                    dictado=dictado,
                    alumno=alumno
                ),
                prefix="int"
            )
        }

    def get(self, request, *args, **kwargs):

        profesor = request.user.persona.perfil_profesor

        dictado = get_object_or_404(
            Dictado,
            pk=kwargs["dictado_id"],
            profesor=profesor
        )

        # Verifica que el alumno pertenezca al dictado
        get_object_or_404(
            InscripcionDictado,
            dictado=dictado,
            alumno_id=kwargs["alumno_id"]
        )

        alumno = get_object_or_404(
            Alumno,
            pk=kwargs["alumno_id"]
        )

        formsets = self.get_formsets(dictado, alumno)

        context = self.get_context_data(**kwargs)
        context.update({
            "formsets": formsets,
            "alumno": alumno,
            "dictado": dictado
        })

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):

        profesor = request.user.persona.perfil_profesor

        dictado = get_object_or_404(
            Dictado,
            pk=kwargs["dictado_id"],
            profesor=profesor
        )

        # Verifica que el alumno pertenezca al dictado
        get_object_or_404(
            InscripcionDictado,
            dictado=dictado,
            alumno_id=kwargs["alumno_id"]
        )

        alumno = get_object_or_404(
            Alumno,
            pk=kwargs["alumno_id"]
        )

        formsets = self.get_formsets(
            dictado,
            alumno,
            request.POST
        )

        if all(fs.is_valid() for fs in formsets.values()):

            formsets["etapa"].save()

            act_forms = formsets["act"].save(commit=False)
            for obj in act_forms:
                obj.dictado = dictado
                obj.alumno = alumno
                obj.save()

            for obj in formsets["act"].deleted_objects:
                obj.delete()

            int_forms = formsets["int"].save(commit=False)
            for obj in int_forms:
                obj.dictado = dictado
                obj.alumno = alumno
                obj.save()

            for obj in formsets["int"].deleted_objects:
                obj.delete()

            messages.success(request, "Calificaciones guardadas correctamente.")

            return redirect("dictado_alumnos", pk=dictado.pk)

        context = self.get_context_data(**kwargs)
        context.update({
            "formsets": formsets,
            "alumno": alumno,
            "dictado": dictado
        })

        return self.render_to_response(context)


    from django.views.generic import TemplateView
from django.db.models import Q

from .models import (
    Curso,
    Especialidad,
    HistorialAcademico,
    CicloLectivo
)

# --------------------------------------------------------------------------------------
# ---                              GestionBoletinesView                              ---
# -------------------------------------------------------------------------------------- 

class GestionBoletinesView(TemplateView):
    template_name = "calificaciones/gestion_boletines.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        q = self.request.GET.get("q", "")
        especialidad = self.request.GET.get("especialidad", "")
        curso_id = self.request.GET.get("curso")

        cursos = Curso.objects.select_related("especialidad").all()

        if q:
            cursos = cursos.filter(
                Q(nivel__icontains=q) |
                Q(division__icontains=q) |
                Q(especialidad__nombre__icontains=q)
            )

        if especialidad:
            cursos = cursos.filter(especialidad_id=especialidad)

        alumnos = []

        if curso_id:
            ciclo = CicloLectivo.objects.get(activo=True)

            alumnos = (
                HistorialAcademico.objects
                .filter(
                    curso_id=curso_id,
                    ciclo_lectivo=ciclo
                )
                .select_related(
                    "alumno__persona"
                )
                .order_by(
                    "alumno__persona__apellido",
                    "alumno__persona__nombre"
                )
            )

        context["cursos"] = cursos
        context["especialidades"] = Especialidad.objects.order_by("nombre")
        context["curso_seleccionado"] = curso_id
        context["especialidad_seleccionada"] = especialidad
        context["q"] = q
        context["alumnos"] = alumnos

        return context

# --------------------------------------------------------------------------------------
# ---                            TablonComunicadosView                              ---
# --------------------------------------------------------------------------------------

class TablonComunicadosView(LoginRequiredMixin, ListView):
    model = Comunicado
    template_name = 'comunicados/lista_comunicados.html'
    context_object_name = 'comunicados'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        filtros = Q(visibilidad=Comunicado.VisibilidadChoices.GLOBAL)

        if user.is_staff or user.user_type in ['PROFESOR', 'PRECEPTOR', 'JERARQUICOS', 'CARGOS']:
            filtros |= Q(visibilidad=Comunicado.VisibilidadChoices.DOCENTES)
            filtros |= Q(visibilidad=Comunicado.VisibilidadChoices.ESTUDIANTES)
        elif user.user_type == 'ALUMNO':
            filtros |= Q(visibilidad=Comunicado.VisibilidadChoices.ESTUDIANTES)

        return Comunicado.objects.filter(filtros)


class CrearComunicadoView(LoginRequiredMixin, CreateView):
    model = Comunicado
    form_class = ComunicadoForm
    template_name = 'comunicados/carga_comunicados.html'

    def get_success_url(self):
        return reverse('lista_comunicados')

    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type == 'ALUMNO' and not request.user.is_staff:
            messages.error(request, "No tenés permisos para publicar comunicados.")
            return redirect('lista_comunicados')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.autor = self.request.user
        messages.success(self.request, "Comunicado publicado de manera exitosa.")
        return super().form_valid(form)

# --------------------------------------------------------------------------------------
# ---                       PlanillaAsistenciaPreceptorView                         ---
# --------------------------------------------------------------------------------------

class PlanillaAsistenciaPreceptorView(TemplateView):
    template_name = 'asistencias/planilla_preceptor.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['ciclos'] = CicloLectivo.objects.all().order_by('-anio')
        context['cursos'] = Curso.objects.all().select_related('especialidad')
        context['turnos'] = Turno.objects.all()

        ciclo_id = self.request.GET.get('ciclo')
        curso_id = self.request.GET.get('curso')
        turno_id = self.request.GET.get('turno')
        burbuja_id = self.request.GET.get('burbuja')
        grupo_taller = self.request.GET.get('grupo_taller')
        fecha_str = self.request.GET.get('fecha')

        fecha_seleccionada = parse_date(fecha_str) if fecha_str else date.today()
        context['selected_fecha'] = fecha_seleccionada.strftime('%Y-%m-%d')

        context['selected_ciclo'] = int(ciclo_id) if ciclo_id else None
        context['selected_curso'] = int(curso_id) if curso_id else None
        context['selected_turno'] = int(turno_id) if turno_id else None
        context['selected_burbuja'] = int(burbuja_id) if burbuja_id else None
        context['selected_grupo_taller'] = grupo_taller if grupo_taller else None

        if context['selected_ciclo']:
            context['burbujas'] = Burbuja.objects.filter(ciclo_lectivo_id=ciclo_id)
        else:
            context['burbujas'] = []

        if ciclo_id and curso_id and turno_id:
            queryset_inscripciones = HistorialAcademico.objects.filter(
                ciclo_lectivo_id=ciclo_id,
                curso_id=curso_id,
                estado_final='CURSANDO'
            ).select_related('alumno__persona')

            if burbuja_id:
                queryset_inscripciones = queryset_inscripciones.filter(burbuja_id=burbuja_id)
            if grupo_taller:
                queryset_inscripciones = queryset_inscripciones.filter(grupo_taller=grupo_taller)

            inscripciones = queryset_inscripciones.order_by('alumno__persona__apellido', 'alumno__persona__nombre')

            asistencias_existentes = Asistencia.objects.filter(
                fecha=fecha_seleccionada,
                turno_id=turno_id,
                inscripcion__in=inscripciones
            ).values_list('inscripcion_id', 'estado', 'burbuja_sesion_id', 'id')

            asistencias_dict = {
                a[0]: {'estado': a[1], 'burbuja_sesion_id': a[2], 'asistencia_id': a[3]}
                for a in asistencias_existentes
            }

            matriz_alumnos = []
            for insc in inscripciones:
                registro = asistencias_dict.get(insc.id)
                matriz_alumnos.append({
                    'inscripcion': insc,
                    'asistencia': registro
                })

            context['matriz_alumnos'] = matriz_alumnos
            context['estados_choices'] = Asistencia.EstadoAsistencia.choices

        return context

    def post(self, request, *args, **kwargs):
        ciclo_id = request.GET.get('ciclo')
        curso_id = request.GET.get('curso')
        turno_id = request.GET.get('turno')
        burbuja_id = request.GET.get('burbuja')
        grupo_taller = request.GET.get('grupo_taller')
        fecha_str = request.GET.get('fecha') or str(date.today())
        fecha_obj = parse_date(fecha_str)

        if not ciclo_id or not curso_id or not turno_id or not fecha_obj:
            messages.error(request, "Error crítico: Faltan parámetros estructurales en el envío del Formulario.")
            return redirect(reverse('planilla_preceptor'))

        try:
            turno_instancia = Turno.objects.get(id=turno_id)
        except Turno.DoesNotExist:
            messages.error(request, "El turno seleccionado es inválido.")
            return redirect(reverse('planilla_preceptor'))

        burbuja_sesion_instancia = Burbuja.objects.filter(id=burbuja_id).first() if burbuja_id else None

        with transaction.atomic():
            for key, value in request.POST.items():
                if key.startswith('asistencia_insc_'):
                    inscripcion_id = key.replace('asistencia_insc_', '')

                    Asistencia.objects.update_or_create(
                        inscripcion_id=inscripcion_id,
                        fecha=fecha_obj,
                        turno=turno_instancia,
                        defaults={
                            'estado': value,
                            'registrado_por': request.user if request.user.is_authenticated else None,
                            'burbuja_sesion': burbuja_sesion_instancia
                        }
                    )

            messages.success(request, f"Parte diario del {fecha_obj.strftime('%d/%m/%Y')} registrado con éxito para este curso.")

        url_retorno = f"{reverse('planilla_preceptor')}?ciclo={ciclo_id}&curso={curso_id}&turno={turno_id}"
        if burbuja_id:
            url_retorno += f"&burbuja={burbuja_id}"
        if grupo_taller:
            url_retorno += f"&grupo_taller={grupo_taller}"
        url_retorno += f"&fecha={fecha_str}"

        return redirect(url_retorno)

# --------------------------------------------------------------------------------------
# ---                            ListaAsistenciasView                               ---
# --------------------------------------------------------------------------------------

class ListaAsistenciasView(LoginRequiredMixin, ListView):
    model = HistorialAcademico
    template_name = 'asistencias/lista_asistencias.html'
    context_object_name = 'inscripciones_control'

    def get_queryset(self):
        queryset = HistorialAcademico.objects.select_related(
            'alumno__persona',
            'curso__especialidad',
            'ciclo_lectivo',
            'burbuja'
        ).order_by('curso', 'alumno__persona__apellido', 'alumno__persona__nombre')

        ciclo_id = self.request.GET.get('ciclo')
        curso_id = self.request.GET.get('curso')
        search_query = self.request.GET.get('q')

        if ciclo_id:
            queryset = queryset.filter(ciclo_lectivo_id=ciclo_id)
        else:
            queryset = queryset.filter(ciclo_lectivo__activo=True)

        if curso_id:
            queryset = queryset.filter(curso_id=curso_id)

        if search_query:
            queryset = queryset.filter(
                Q(alumno__persona__nombre__icontains=search_query) |
                Q(alumno__persona__apellido__icontains=search_query) |
                Q(alumno__persona__dni__icontains=search_query)
            )

        queryset = queryset.annotate(
            total_presentes=Count(
                Case(When(asistencias__estado='P', then=1))
            ),
            total_justificadas=Count(
                Case(When(asistencias__estado='J', then=1))
            ),
            total_faltas=Sum(
                Case(
                    When(asistencias__estado='A', then=F('asistencias__turno__valor_falta')),
                    default=0,
                    output_field=DecimalField()
                )
            )
        )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['ciclos'] = CicloLectivo.objects.all().order_by('-anio')
        context['cursos'] = Curso.objects.all()

        context['selected_ciclo'] = int(self.request.GET.get('ciclo')) if self.request.GET.get('ciclo') else None
        context['selected_curso'] = int(self.request.GET.get('curso')) if self.request.GET.get('curso') else None
        context['search_query'] = self.request.GET.get('q', '')

        return context

# --------------------------------------------------------------------------------------
# ---                     DetalleAsistenciasAlumnoView                               ---
# --------------------------------------------------------------------------------------

class DetalleAsistenciasAlumnoView(LoginRequiredMixin, DetailView):
    model = Alumno
    template_name = 'asistencias/detalle_asistencias_alumno.html'
    context_object_name = 'alumno'

    def get_queryset(self):
        return Alumno.objects.select_related('persona').prefetch_related(
            Prefetch(
                'historiales',
                queryset=HistorialAcademico.objects.select_related(
                    'curso__especialidad',
                    'ciclo_lectivo',
                    'burbuja'
                ).prefetch_related(
                    'asistencias__turno'
                ).order_by('-ciclo_lectivo__anio'),
                to_attr='historial_asistencias'
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        historial_agrupado = {}

        for historial in self.object.historial_asistencias:
            anio = historial.ciclo_lectivo.anio

            lista_asistencias = historial.asistencias.all().order_by('-fecha', 'turno__nombre')

            presentes = sum(1 for a in lista_asistencias if a.estado == 'P')
            justificadas = sum(1 for a in lista_asistencias if a.estado == 'J')

            faltas_totales = sum(
                float(a.turno.valor_falta) for a in lista_asistencias if a.estado == 'A'
            )

            historial_agrupado[anio] = {
                'curso': historial.curso,
                'burbuja': historial.burbuja,
                'grupo_taller': historial.get_grupo_taller_display(),
                'estado_final': historial.get_estado_final_display(),
                'totales': {
                    'presentes': presentes,
                    'justificadas': justificadas,
                    'faltas': faltas_totales
                },
                'registros': lista_asistencias
            }

        context['historial_asistencias_agrupado'] = historial_agrupado
        return context

# --------------------------------------------------------------------------------------
# ---                           PlanillaCargaNotasView                               ---
# --------------------------------------------------------------------------------------

class PlanillaCargaNotasView(LoginRequiredMixin, TemplateView):
    template_name = 'calificaciones/planilla_docente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['ciclos'] = CicloLectivo.objects.all().order_by('-anio')
        context['cursos'] = Curso.objects.all()

        ciclo_id = self.request.GET.get('ciclo')
        curso_id = self.request.GET.get('curso')
        dictado_id = self.request.GET.get('dictado')

        context['selected_ciclo'] = int(ciclo_id) if ciclo_id else None
        context['selected_curso'] = int(curso_id) if curso_id else None
        context['selected_dictado'] = int(dictado_id) if dictado_id else None

        if context['selected_curso']:
            context['dictados'] = Dictado.objects.filter(curso_id=curso_id).select_related('materia')
        else:
            context['dictados'] = []

        if ciclo_id and curso_id and dictado_id:
            inscripciones = InscripcionDictado.objects.filter(
                dictado_id=dictado_id,
                ciclo_lectivo_id=ciclo_id
            ).select_related('alumno__persona').order_by('alumno__persona__apellido', 'alumno__persona__nombre')

            notas_existentes = NotaActividad.objects.filter(dictado_id=dictado_id).order_by('fecha', 'id')
            etapas_existentes = NotaEtapa.objects.filter(dictado_id=dictado_id).order_by('etapa')
            intensificaciones_existentes = Intensificacion.objects.filter(dictado_id=dictado_id).order_by('fecha', 'id')

            actividades_cabecera = []
            seen_actividades = set()
            for n in notas_existentes:
                if n.nombre_actividad not in seen_actividades:
                    actividades_cabecera.append({'nombre': n.nombre_actividad, 'cuatrimestre': n.cuatrimestre})
                    seen_actividades.add(n.nombre_actividad)

            etapas_cabecera = []
            seen_etapas = set()
            for e in etapas_existentes:
                if e.etapa not in seen_etapas:
                    etapas_cabecera.append(e.etapa)
                    seen_etapas.add(e.etapa)

            conteo_por_alumno_fecha = {}
            for i in intensificaciones_existentes:
                key = (i.alumno_id, i.fecha)
                conteo_por_alumno_fecha[key] = conteo_por_alumno_fecha.get(key, 0) + 1

            max_subcolumnas_por_fecha = {}
            for (alumno_id, fch), cantidad in conteo_por_alumno_fecha.items():
                if cantidad > max_subcolumnas_por_fecha.get(fch, 0):
                    max_subcolumnas_por_fecha[fch] = cantidad

            intensificaciones_cabecera = []
            fechas_unicas_ordenadas = sorted(list(set(i.fecha for i in intensificaciones_existentes if i.fecha)))

            for fch in fechas_unicas_ordenadas:
                total_columnas_ese_dia = max_subcolumnas_por_fecha.get(fch, 1)
                for index in range(total_columnas_ese_dia):
                    intensificaciones_cabecera.append({
                        'fecha_obj': fch,
                        'str_cabecera': fch.strftime('%d/%m/%Y'),
                        'indice': index
                    })

            context['actividades_cabecera'] = actividades_cabecera
            context['etapas_cabecera'] = etapas_cabecera
            context['intensificaciones_cabecera'] = intensificaciones_cabecera
            context['today'] = date.today()

            matriz_alumnos = []
            for ins in inscripciones:
                alumno = ins.alumno

                columnas_actividades = []
                for act in actividades_cabecera:
                    nota_obj = notas_existentes.filter(alumno_id=alumno.id, nombre_actividad=act['nombre']).first()
                    columnas_actividades.append({
                        'id': nota_obj.id if nota_obj else None,
                        'valor': nota_obj.valor if nota_obj else '',
                        'identificador_vacio': f"{alumno.id}__act__{act['nombre'].replace(' ', '_')}"
                    })

                columnas_etapas = []
                for etapa_nombre in etapas_cabecera:
                    etapa_obj = etapas_existentes.filter(alumno_id=alumno.id, etapa=etapa_nombre).first()
                    columnas_etapas.append({
                        'id': etapa_obj.id if etapa_obj else None,
                        'valor_numerico': etapa_obj.valor_numerico if etapa_obj else '',
                        'identificador_vacio': f"{alumno.id}__etapa__{etapa_nombre.replace(' ', '_')}"
                    })

                columnas_intensificaciones = []
                for cabecera in intensificaciones_cabecera:
                    fch = cabecera['fecha_obj']
                    idx = cabecera['indice']

                    notas_alumno_fecha = intensificaciones_existentes.filter(alumno_id=alumno.id, fecha=fch)

                    inte_obj = None
                    if idx < len(notas_alumno_fecha):
                        inte_obj = notas_alumno_fecha[idx]

                    columnas_intensificaciones.append({
                        'id': inte_obj.id if inte_obj else None,
                        'valor': int(inte_obj.valor) if inte_obj and inte_obj.valor is not None else '',
                        'identificador_vacio': f"{alumno.id}__inte__{fch.strftime('%Y-%m-%d')}__idx__{idx}"
                    })

                matriz_alumnos.append({
                    'alumno': alumno,
                    'notas_columnas': columnas_actividades,
                    'etapas_columnas': columnas_etapas,
                    'intensificaciones_columnas': columnas_intensificaciones
                })

            context['matriz_alumnos'] = matriz_alumnos

        return context

    def post(self, request, *args, **kwargs):
        ciclo_id = request.GET.get('ciclo')
        curso_id = request.GET.get('curso')
        dictado_id = request.GET.get('dictado')

        if not dictado_id:
            return redirect(reverse('planilla_docente'))

        nueva_actividad_nombre = request.POST.get('nuevo_nombre_actividad')
        nuevo_cuatrimestre = request.POST.get('nuevo_cuatrimestre')
        nueva_fecha_intensificacion_str = request.POST.get('nueva_fecha_intensificacion')

        with transaction.atomic():
            for key, value in request.POST.items():
                if value.strip() == '':
                    continue

                if key.startswith('nota_existente_'):
                    valor = _valor_decimal(value)
                    if valor is not None:
                        NotaActividad.objects.filter(id=key.replace('nota_existente_', '')).update(valor=valor)

                elif key.startswith('nota_nueva_celda_'):
                    info = key.replace('nota_nueva_celda_', '').split('__act__')
                    alumno_id, act_nombre = info[0], info[1].replace('_', ' ')
                    act_prev = NotaActividad.objects.filter(dictado_id=dictado_id, nombre_actividad=act_nombre).first()
                    valor = _valor_decimal(value)
                    if valor is not None:
                        NotaActividad.objects.create(
                            dictado_id=dictado_id, alumno_id=alumno_id, nombre_actividad=act_nombre,
                            cuatrimestre=act_prev.cuatrimestre if act_prev else 1, valor=valor,
                            fecha=act_prev.fecha if act_prev else date.today()
                        )

                elif key.startswith('etapa_existente_'):
                    valor = _valor_decimal(value)
                    if valor is not None:
                        NotaEtapa.objects.filter(id=key.replace('etapa_existente_', '')).update(valor_numerico=valor)

                elif key.startswith('etapa_nueva_celda_'):
                    info = key.replace('etapa_nueva_celda_', '').split('__etapa__')
                    valor = _valor_decimal(value)
                    if valor is not None:
                        NotaEtapa.objects.create(dictado_id=dictado_id, alumno_id=info[0], etapa=info[1].replace('_', ' '), valor_numerico=valor)

                elif key.startswith('intensificacion_existente_'):
                    inte_id = key.replace('intensificacion_existente_', '')
                    valor = _valor_decimal(value)
                    if valor is not None:
                        Intensificacion.objects.filter(id=inte_id).update(valor=valor)

                elif key.startswith('intensificacion_nueva_celda_'):
                    info = key.replace('intensificacion_nueva_celda_', '').split('__inte__')
                    alumno_id = info[0]
                    fecha_columna = parse_date(info[1].split('__idx__')[0])
                    valor = _valor_decimal(value)
                    if valor is not None:
                        Intensificacion.objects.create(
                            dictado_id=dictado_id, alumno_id=alumno_id,
                            valor=valor, fecha=fecha_columna
                        )

            if nueva_actividad_nombre and nueva_actividad_nombre.strip():
                nombre_limpio = nueva_actividad_nombre.strip()
                fecha_carga = parse_date(request.POST.get('nueva_fecha_actividad')) or date.today()
                for k, val in request.POST.items():
                    if k.startswith('nueva_nota_alumno_') and val.strip() != '':
                        valor = _valor_decimal(val)
                        if valor is not None:
                            NotaActividad.objects.create(
                                dictado_id=dictado_id, alumno_id=k.replace('nueva_nota_alumno_', ''),
                                nombre_actividad=nombre_limpio, cuatrimestre=int(nuevo_cuatrimestre),
                                valor=valor, fecha=fecha_carga
                            )

            if nueva_fecha_intensificacion_str:
                fecha_int_nueva = parse_date(nueva_fecha_intensificacion_str)
                if fecha_int_nueva:
                    contador_altas = 0
                    for k, val in request.POST.items():
                        if k.startswith('nueva_intensificacion_alumno_') and val.strip() != '':
                            alumno_id = k.replace('nueva_intensificacion_alumno_', '')
                            valor = _valor_decimal(val)
                            if valor is not None:
                                Intensificacion.objects.create(
                                    dictado_id=dictado_id,
                                    alumno_id=alumno_id,
                                    valor=valor,
                                    fecha=fecha_int_nueva
                                )
                                contador_altas += 1
                    if contador_altas > 0:
                        messages.success(request, f"Nueva columna de intensificación registrada para el día {fecha_int_nueva.strftime('%d/%m/%Y')}.")

        return redirect(f"{reverse('planilla_docente')}?ciclo={ciclo_id}&curso={curso_id}&dictado={dictado_id}")

# --------------------------------------------------------------------------------------
# ---                         CargaFormsetAlumnosView                                ---
# --------------------------------------------------------------------------------------

class CargaFormsetAlumnosView(LoginRequiredMixin, TemplateView):
    template_name = 'alumnos/carga_dinamica_alumnos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ciclos'] = CicloLectivo.objects.filter(activo=True)
        context['cursos'] = Curso.objects.all()
        context['formset'] = AlumnoFormSet(prefix='alumnos')
        return context

    def post(self, request, *args, **kwargs):
        ciclo_id = request.POST.get('ciclo_lectivo')
        curso_id = request.POST.get('curso')

        if not ciclo_id or not curso_id:
            messages.error(request, "Falta seleccionar el Ciclo Lectivo o el Curso de destino.")
            return self.render_to_response(self.get_context_data())

        try:
            ciclo = CicloLectivo.objects.get(id=ciclo_id)
            curso_destino = Curso.objects.get(id=curso_id)
        except (CicloLectivo.DoesNotExist, Curso.DoesNotExist):
            messages.error(request, "El Ciclo Lectivo o el Curso seleccionado no existe.")
            return self.render_to_response(self.get_context_data())

        formset = AlumnoFormSet(request.POST, prefix='alumnos')

        if formset.is_valid():
            dictados_curso = Dictado.objects.filter(curso=curso_destino, ciclo_lectivo=ciclo)

            if not dictados_curso.exists():
                messages.error(request, f"El curso {curso_destino} no posee materias (Dictados) asociadas en este ciclo lectivo.")
                return self.render_to_response(self.get_context_data())

            creados = 0
            errores = []

            try:
                with transaction.atomic():
                    for form in formset:
                        if not form.cleaned_data or not form.cleaned_data.get('dni'):
                            continue

                        dni = form.cleaned_data['dni'].strip()
                        apellido = form.cleaned_data['apellido'].strip()
                        nombre = form.cleaned_data['nombre'].strip()

                        legajo_in = form.cleaned_data.get('numero_legajo')
                        libro_in = form.cleaned_data.get('libro')
                        folio_in = form.cleaned_data.get('folio')

                        if Persona.objects.filter(dni=dni).exists():
                            errores.append(f"El DNI {dni} ya está registrado en el sistema.")
                            continue

                        legajo_final = legajo_in.strip() if legajo_in else f"LEG-{dni}-{random.randint(10,99)}"
                        if Persona.objects.filter(numero_legajo=legajo_final).exists():
                            legajo_final = f"LEG-{dni}-{random.randint(100,999)}"

                        cuil_dummy = f"20{dni}7"
                        email_dummy = f"alu.{dni}@institucion.edu.ar"

                        username = f"alu_{dni}"
                        user = User.objects.create_user(
                            username=username,
                            password=dni,
                            first_name=nombre[:30],
                            last_name=apellido[:30]
                        )

                        persona = Persona.objects.create(
                            user=user,
                            dni=dni,
                            cuil=cuil_dummy,
                            apellido=apellido,
                            nombre=nombre,
                            numero_legajo=legajo_final,
                            email=email_dummy,
                            telefono="S/D",
                            domicilio="S/D"
                        )

                        libro_final = libro_in.strip() if libro_in else "S/D"
                        folio_final = folio_in.strip() if folio_in else "S/D"

                        alumno = Alumno.objects.create(
                            persona=persona,
                            libro=libro_final,
                            folio=folio_final,
                            activo=True
                        )

                        HistorialAcademico.objects.create(
                            alumno=alumno,
                            curso=curso_destino,
                            ciclo_lectivo=ciclo,
                            estado_final='CURSANDO'
                        )

                        for dictado in dictados_curso:
                            InscripcionDictado.objects.create(
                                alumno=alumno,
                                dictado=dictado,
                                ciclo_lectivo=ciclo,
                                condicion='REGULAR'
                            )

                        creados += 1

            except Exception as e:
                messages.error(request, f"Error de integridad en base de datos: {str(e)}")
                return self.render_to_response(self.get_context_data())

            if creados > 0:
                messages.success(request, f"Se cargaron {creados} alumnos con éxito en {curso_destino} y se les generó la inscripción automática a {dictados_curso.count()} materias.")

            if errores:
                messages.warning(request, " ".join(errores))
        else:
            messages.error(request, "Revisá los datos cargados en la nómina. Verificá que el DNI, Apellido y Nombre estén completos en cada fila.")

        return redirect(reverse('carga_dinamica_alumnos'))

# --------------------------------------------------------------------------------------
# ---                           AltaPersonalCargoView                                ---
# --------------------------------------------------------------------------------------

class AltaPersonalCargoView(LoginRequiredMixin, FormView):
    template_name = 'cargos/alta_cargos.html'
    form_class = AltaPersonalCargoForm
    success_url = reverse_lazy('lista_cargos')

    def form_valid(self, form):
        form.save()
        messages.success(self.request, "Personal registrado y cargo asignado de manera exitosa.")
        return super().form_valid(form)